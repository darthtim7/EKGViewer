#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
import re
import sqlite3
import sys
import tempfile
import zipfile
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader

RESTORE_BYTES = 211_294_688
RESTORE_SHA256 = "cbc08d30997f3f9b9516100a6ed1e75dd6376e3e23456bfba2beee09e35b666d"
PROJECT_BYTES = 211_898_622
PROJECT_SHA256 = "9e1f0ce6f3f26784fb2b378b275adcc416d36883ca94a1a39e643252f17b6216"
PUBLICATION_SHA256 = "8a053112ca24cd730b970130d5d0fc57a15c681531603601096186aeb0cd9642"
APPLICATION_SHA256 = "5f1e4ac8fc6e2ffad213646c78e4f261bf655795de5ac8a7d4486d3be11ce139"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def verify_zip(path: Path) -> dict[str, Any]:
    with zipfile.ZipFile(path) as zf:
        bad = zf.testzip()
        names = zf.namelist()
        unsafe = []
        for name in names:
            pp = PurePosixPath(name.replace("\\", "/"))
            if pp.is_absolute() or ".." in pp.parts or re.match(r"^[A-Za-z]:", name):
                unsafe.append(name)
    result = {
        "path": str(path),
        "bytes": path.stat().st_size,
        "sha256": sha256_file(path),
        "members": len(names),
        "crc_error": bad,
        "duplicates": len(names) - len(set(names)),
        "unsafe_paths": unsafe,
    }
    if bad or result["duplicates"] or unsafe:
        raise RuntimeError({"zip_verification_failed": result})
    return result


def safe_extract(path: Path, destination: Path) -> None:
    verify_zip(path)
    destination.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path) as zf:
        zf.extractall(destination)


def find_unique_by_identity(root: Path, *, size: int, digest: str) -> Path:
    matches = []
    for path in root.rglob("*"):
        if not path.is_file() or path.stat().st_size != size:
            continue
        if sha256_file(path) == digest:
            matches.append(path)
    if len(matches) != 1:
        raise RuntimeError({"identity_candidates": [str(path) for path in matches], "bytes": size, "sha256": digest})
    return matches[0]


def reconstruct_restore(inputs: Path, work: Path) -> tuple[Path, dict[str, Any]]:
    wrappers = sorted(path for path in inputs.rglob("*.zip") if path.is_file())
    if len(wrappers) < 3:
        raise RuntimeError({"wrapper_candidates": [str(path) for path in wrappers]})
    extracted = work / "volume_wrappers"
    extracted.mkdir(parents=True)
    wrapper_qa = []
    for index, wrapper in enumerate(wrappers, start=1):
        wrapper_qa.append(verify_zip(wrapper))
        safe_extract(wrapper, extracted / f"wrapper-{index:02d}")
    part_map: dict[str, list[Path]] = {"part001": [], "part002": [], "part003": []}
    for path in extracted.rglob("*"):
        if not path.is_file():
            continue
        lower = path.name.lower()
        for suffix in part_map:
            if lower.endswith(suffix) or f".{suffix}" in lower:
                part_map[suffix].append(path)
    selected = []
    for suffix in ("part001", "part002", "part003"):
        candidates = part_map[suffix]
        unique: dict[tuple[int, str], Path] = {}
        for path in candidates:
            unique[(path.stat().st_size, sha256_file(path))] = path
        if len(unique) != 1:
            raise RuntimeError({suffix: [{"path": str(p), "bytes": p.stat().st_size, "sha256": sha256_file(p)} for p in candidates]})
        selected.append(next(iter(unique.values())))
    restore = work / "response77-complete-restore.zip"
    with restore.open("wb") as output:
        for part in selected:
            with part.open("rb") as source:
                for block in iter(lambda: source.read(1024 * 1024), b""):
                    output.write(block)
    if restore.stat().st_size != RESTORE_BYTES or sha256_file(restore) != RESTORE_SHA256:
        raise RuntimeError({"restore_identity": {"bytes": restore.stat().st_size, "sha256": sha256_file(restore)}})
    restore_qa = verify_zip(restore)
    return restore, {"wrappers": wrapper_qa, "parts": [{"path": str(p), "bytes": p.stat().st_size, "sha256": sha256_file(p)} for p in selected], "restore": restore_qa}


def locate_project_root(extracted: Path) -> Path:
    roots = [path for path in extracted.iterdir() if path.is_dir()]
    if len(roots) == 1:
        return roots[0]
    return extracted


def inspect_database(path: Path) -> dict[str, Any]:
    con = sqlite3.connect(path)
    try:
        integrity = con.execute("PRAGMA integrity_check").fetchone()[0]
        fk = list(con.execute("PRAGMA foreign_key_check"))
        tables = [row[0] for row in con.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")]
        result: dict[str, Any] = {
            "path": str(path),
            "bytes": path.stat().st_size,
            "sha256": sha256_file(path),
            "integrity": integrity,
            "foreign_key_violations": len(fk),
            "table_count": len(tables),
            "tables": tables,
        }
        for table in (
            "thread_response_reconciliation_cp3",
            "fractional_prompt_cp3",
            "section5_session1_checkpoint",
            "section5_print_selection",
            "section5_print_derivative",
            "section5_page_transform",
            "section5_session1_release",
            "section5_recovery_event",
            "recovery_event_cp3",
        ):
            if table not in tables:
                continue
            columns = [row[1] for row in con.execute(f"PRAGMA table_info({table})")]
            count = con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            samples = []
            for row in con.execute(f"SELECT * FROM {table} LIMIT 5"):
                samples.append(dict(zip(columns, row)))
            result.setdefault("relevant_tables", {})[table] = {"columns": columns, "count": count, "samples": samples}
        if "thread_response_reconciliation_cp3" in tables:
            columns = [row[1] for row in con.execute("PRAGMA table_info(thread_response_reconciliation_cp3)")]
            if "response_key" in columns:
                result["response77_records"] = con.execute("SELECT COUNT(*) FROM thread_response_reconciliation_cp3 WHERE response_key='R77'").fetchone()[0]
                result["maximum_response_number"] = con.execute("SELECT MAX(CAST(response_number AS INTEGER)) FROM thread_response_reconciliation_cp3").fetchone()[0]
        return result
    finally:
        con.close()


def inspect_project(project: Path) -> dict[str, Any]:
    files = [path for path in project.rglob("*") if path.is_file()]
    databases = []
    for path in project.rglob("*.sqlite"):
        try:
            item = inspect_database(path)
        except Exception as exc:
            item = {"path": str(path), "error": repr(exc)}
        databases.append(item)
    databases.sort(key=lambda row: (row.get("response77_records", 0), row.get("table_count", 0), row.get("bytes", 0)), reverse=True)

    workbooks = []
    for path in project.rglob("*.xlsx"):
        try:
            wb = load_workbook(path, read_only=True, data_only=False)
            try:
                sheet_names = list(wb.sheetnames)
            finally:
                wb.close()
            workbooks.append({"path": str(path), "bytes": path.stat().st_size, "sha256": sha256_file(path), "sheet_count": len(sheet_names), "sheet_names": sheet_names})
        except Exception as exc:
            workbooks.append({"path": str(path), "error": repr(exc)})
    workbooks.sort(key=lambda row: (row.get("sheet_count", 0), row.get("bytes", 0)), reverse=True)

    pdfs = []
    for path in project.rglob("*.pdf"):
        try:
            digest = sha256_file(path)
            reader = PdfReader(str(path))
            page_count = len(reader.pages)
            searchable = None
            if page_count in (537, 538):
                searchable = sum(1 for page in reader.pages if (page.extract_text() or "").strip())
            pdfs.append({"path": str(path), "bytes": path.stat().st_size, "sha256": digest, "page_count": page_count, "searchable_pages": searchable})
        except Exception as exc:
            pdfs.append({"path": str(path), "error": repr(exc)})
    pdfs.sort(key=lambda row: (row.get("page_count", 0), row.get("bytes", 0)), reverse=True)

    pngs = []
    for path in project.rglob("*.png"):
        try:
            with Image.open(path) as image:
                width, height = image.size
                mode = image.mode
            if width >= 5000 or height >= 3000:
                pngs.append({"path": str(path), "bytes": path.stat().st_size, "sha256": sha256_file(path), "pixels": [width, height], "mode": mode})
        except Exception:
            pass

    apps = []
    for path in project.rglob("human_pathogen_app.py"):
        if path.is_file():
            apps.append({"path": str(path), "bytes": path.stat().st_size, "sha256": sha256_file(path), "matches_governed": sha256_file(path) == APPLICATION_SHA256})

    manifests = [str(path.relative_to(project)) for path in project.rglob("*manifest*") if path.is_file()]
    readmes = [str(path.relative_to(project)) for path in project.rglob("README*") if path.is_file()]
    return {
        "project_root": str(project),
        "physical_file_count": len(files),
        "physical_bytes": sum(path.stat().st_size for path in files),
        "database_candidates": databases[:20],
        "workbook_candidates": workbooks[:20],
        "pdf_candidates": pdfs[:40],
        "large_png_candidates": pngs,
        "application_candidates": apps,
        "manifest_paths": manifests[:100],
        "readme_paths": readmes[:100],
        "publication_candidates": [row for row in pdfs if row.get("sha256") == PUBLICATION_SHA256],
    }


def main() -> None:
    inputs = Path(sys.argv[1] if len(sys.argv) > 1 else "inputs/response77")
    output = Path(sys.argv[2] if len(sys.argv) > 2 else "inspection_response77")
    output.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="mrhpd-r77-inspect-") as td:
        work = Path(td)
        restore, transport = reconstruct_restore(inputs, work)
        restore_root = work / "restore"
        safe_extract(restore, restore_root)
        project_zip = find_unique_by_identity(restore_root, size=PROJECT_BYTES, digest=PROJECT_SHA256)
        project_qa = verify_zip(project_zip)
        extracted_project = work / "project"
        safe_extract(project_zip, extracted_project)
        project_root = locate_project_root(extracted_project)
        project = inspect_project(project_root)
        result = {
            "schema": "mrhpd-response77-project-inspection-1.0",
            "status": "passed",
            "transport": transport,
            "project_archive": project_qa,
            "project": project,
        }
        path = output / "MRHPD_RESPONSE77_PROJECT_INSPECTION.json"
        path.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
