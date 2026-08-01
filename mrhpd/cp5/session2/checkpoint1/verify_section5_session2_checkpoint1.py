#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
import re
import sys
import tempfile
import zipfile
from pathlib import Path, PurePosixPath

from openpyxl import load_workbook
from pypdf import PdfReader


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def verify_zip(path: Path) -> list[str]:
    with zipfile.ZipFile(path) as zf:
        if zf.testzip() is not None:
            raise RuntimeError(f"ZIP CRC failure: {path}")
        names = zf.namelist()
        if len(names) != len(set(names)):
            raise RuntimeError(f"duplicate ZIP members: {path}")
        for name in names:
            pp = PurePosixPath(name.replace("\\", "/"))
            if pp.is_absolute() or ".." in pp.parts or re.match(r"^[A-Za-z]:", name):
                raise RuntimeError(f"unsafe ZIP member: {name}")
            if any(token in name.lower() for token in ("filler", "padding", "dummy_payload", "artificial_inflation")):
                raise RuntimeError(f"artificial filler member: {name}")
    return names


def safe_extract(path: Path, destination: Path) -> None:
    verify_zip(path)
    destination.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path) as zf:
        zf.extractall(destination)


def main() -> None:
    dist = Path(sys.argv[1] if len(sys.argv) > 1 else "dist_cp5_s2_cp1")
    deliveries = list(dist.glob("MRHPD v3.0.0a Response 79 Section 5 Session 2 Checkpoint 1 Recovery Package *.zip"))
    if len(deliveries) != 1:
        raise RuntimeError({"delivery_candidates": [str(path) for path in deliveries]})
    delivery = deliveries[0]
    names = verify_zip(delivery)
    required_tokens = [
        "Recovery Verification.json",
        "BUILD_SUMMARY.json",
        "Exact File Names.txt",
        ".sha256.txt",
        "Provider Preview and Physical Proof Intake Report.pdf",
        "Provider Preview and Physical Proof Intake Report.docx",
        "Provider Preview and Proof Register.xlsx",
        "Print Previewer Evidence Register.csv",
        "Provider Conversion Issue Register.csv",
        "Physical Proof Plan.csv",
        "Physical Proof Inspection Register.csv",
    ]
    missing = [token for token in required_tokens if not any(token in name for name in names)]
    if missing:
        raise RuntimeError({"missing_delivery_controls": missing})
    summary_path = dist / "MRHPD_RESPONSE79_SECTION5_SESSION2_CHECKPOINT1_BUILD_SUMMARY.json"
    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    with tempfile.TemporaryDirectory(prefix="mrhpd-r79-independent-") as td:
        extracted = Path(td)
        safe_extract(delivery, extracted)
        report_pdfs = list(extracted.glob("*Provider Preview and Physical Proof Intake Report.pdf"))
        report_xlsxs = list(extracted.glob("*Provider Preview and Proof Register.xlsx"))
        if len(report_pdfs) != 1 or len(report_xlsxs) != 1:
            raise RuntimeError({"report_pdf_candidates": [str(path) for path in report_pdfs], "report_register_candidates": [str(path) for path in report_xlsxs]})
        report_pdf = report_pdfs[0]
        report_xlsx = report_xlsxs[0]
        reader = PdfReader(str(report_pdf))
        searchable = sum(1 for page in reader.pages if (page.extract_text() or "").strip())
        wb = load_workbook(report_xlsx, read_only=True, data_only=False)
        try:
            register_sheets = len(wb.sheetnames)
        finally:
            wb.close()
    preview_state = summary.get("provider_previewer", {})
    preview_count = preview_state.get("preview_evidence_records", preview_state.get("evidence_records"))
    gates = {
        "summary": summary.get("status") == "passed",
        "clean_apply": summary.get("recovery", {}).get("clean_apply", {}).get("status") == "passed",
        "database": summary.get("database", {}).get("integrity") == "ok" and summary.get("database", {}).get("foreign_key_violations") == 0,
        "responses": summary.get("database", {}).get("response78_records") == 1 and summary.get("database", {}).get("response79_records") == 1,
        "workbook": summary.get("workbook", {}).get("current_sheet_count", 0) >= 114 and summary.get("workbook", {}).get("formula_error_count") == 0,
        "preview_evidence": preview_count == 8 and preview_state.get("unsupported_approval_claims") == 0,
        "proof_plan": summary.get("physical_proof", {}).get("plan_records") == 11 and summary.get("physical_proof", {}).get("completed_inspections") == 0,
        "digital_publication": summary.get("digital_publication", {}).get("sha256") == "8a053112ca24cd730b970130d5d0fc57a15c681531603601096186aeb0cd9642" and summary.get("digital_publication", {}).get("pages") == 537,
        "print_interior": summary.get("print_interior", {}).get("sha256") == "0216def4f41b2b62fc2eb3f87f5a66abbf633e54c41b31e2b39afa29c34b0803" and summary.get("print_interior", {}).get("pages") == 538,
        "cover": summary.get("cover", {}).get("sha256") == "3945225ef87c87a8795354aee1c90ce58d39fd6d5bb57229489692420ba07097",
        "application": summary.get("application", {}).get("main_application_unchanged") is True,
        "report_pdf": len(reader.pages) >= 3 and searchable == len(reader.pages),
        "report_register": register_sheets == 8,
        "user_upload": summary.get("user_upload_required") is False,
        "checkpoint": summary.get("checkpoint_1_of_3_complete") is True and summary.get("session_2_of_3_complete") is False,
    }
    if not all(gates.values()):
        raise RuntimeError({"independent_output_gate_failed": gates})
    result = {
        "status": "passed",
        "delivery": delivery.name,
        "bytes": delivery.stat().st_size,
        "sha256": sha256_file(delivery),
        "members": len(names),
        "recovery_zip": summary["recovery"]["recovery_zip"],
        "database_tables": summary["database"]["table_count"],
        "workbook_sheets": summary["workbook"]["current_sheet_count"],
        "preview_evidence_records": preview_count,
        "proof_plan_records": summary["physical_proof"]["plan_records"],
        "provider_approval_claimed": False,
        "physical_proof_claimed": False,
        "user_upload_required": False,
    }
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
