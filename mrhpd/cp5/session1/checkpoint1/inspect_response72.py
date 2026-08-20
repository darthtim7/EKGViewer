#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
import sys
import tempfile
import zipfile
from pathlib import Path, PurePosixPath
from typing import Any

RESTORE_BYTES = 159_186_352
RESTORE_SHA256 = "cb6d2de9bb351a4ff580e8ac0ac071a774670974098da88be822d64b437b25ce"
PROJECT_BYTES = 159_865_032
PROJECT_SHA256 = "88b3a6fab6e1106b2942b92fbe5b10c9b06ffe6f15963a7f0c308203dcb6beb5"


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def safe_members(zf: zipfile.ZipFile) -> list[zipfile.ZipInfo]:
    names: list[str] = []
    for info in zf.infolist():
        name = info.filename.replace("\\", "/")
        pp = PurePosixPath(name)
        if pp.is_absolute() or ".." in pp.parts or re.match(r"^[A-Za-z]:", name):
            raise RuntimeError(f"unsafe ZIP member: {info.filename}")
        names.append(name)
    if len(names) != len(set(names)):
        raise RuntimeError("duplicate ZIP members")
    if zf.testzip() is not None:
        raise RuntimeError("ZIP CRC failure")
    return zf.infolist()


def safe_extract(path: Path, destination: Path) -> None:
    destination.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path) as zf:
        safe_members(zf)
        zf.extractall(destination)


def identify_wrapper(directory: Path, sequence: int) -> Path:
    candidates = []
    for path in directory.rglob("*.zip"):
        try:
            with zipfile.ZipFile(path) as zf:
                names = zf.namelist()
                if any(name.endswith(f".part{sequence:03d}") for name in names):
                    candidates.append(path)
        except zipfile.BadZipFile:
            continue
    if len(candidates) != 1:
        raise RuntimeError({"wrapper_sequence": sequence, "candidates": [str(p) for p in candidates]})
    return candidates[0]


def reconstruct_restore(volume1_dir: Path, volume2_dir: Path, work: Path) -> Path:
    outer1 = next(volume1_dir.rglob("*.zip"))
    outer2 = next(volume2_dir.rglob("*.zip"))
    e1 = work / "outer1"
    e2 = work / "outer2"
    safe_extract(outer1, e1)
    safe_extract(outer2, e2)
    w1 = identify_wrapper(e1, 1)
    w2 = identify_wrapper(e2, 2)
    x1 = work / "volume1"
    x2 = work / "volume2"
    safe_extract(w1, x1)
    safe_extract(w2, x2)
    part1 = next(x1.rglob("*.part001"))
    part2 = next(x2.rglob("*.part002"))
    manifest_paths = list(x1.rglob("*TRANSPORT_MANIFEST.json")) + list(x2.rglob("*TRANSPORT_MANIFEST.json"))
    if not manifest_paths:
        raise RuntimeError("transport manifest absent")
    manifest = json.loads(manifest_paths[0].read_text(encoding="utf-8-sig"))
    output = work / manifest["restore"]["name"]
    with output.open("wb") as dst:
        for part in (part1, part2):
            with part.open("rb") as src:
                for block in iter(lambda: src.read(1024 * 1024), b""):
                    dst.write(block)
    if output.stat().st_size != RESTORE_BYTES or sha256_file(output) != RESTORE_SHA256:
        raise RuntimeError({"restore_identity": {"bytes": output.stat().st_size, "sha256": sha256_file(output)}})
    with zipfile.ZipFile(output) as zf:
        safe_members(zf)
    return output


def locate_project_archive(restore_root: Path) -> Path:
    candidates = [p for p in restore_root.rglob("*.zip") if "COMPLETE PROJECT THROUGH RESPONSE 72" in p.name]
    if len(candidates) != 1:
        raise RuntimeError({"project_archive_candidates": [str(p) for p in candidates]})
    archive = candidates[0]
    if archive.stat().st_size != PROJECT_BYTES or sha256_file(archive) != PROJECT_SHA256:
        raise RuntimeError({"project_identity": {"bytes": archive.stat().st_size, "sha256": sha256_file(archive)}})
    with zipfile.ZipFile(archive) as zf:
        safe_members(zf)
    return archive


def table_schema(con: sqlite3.Connection, name: str) -> dict[str, Any]:
    cols = [
        {"cid": row[0], "name": row[1], "type": row[2], "notnull": row[3], "default": row[4], "pk": row[5]}
        for row in con.execute(f'PRAGMA table_info("{name}")')
    ]
    count = con.execute(f'SELECT COUNT(*) FROM "{name}"').fetchone()[0]
    sample = []
    if count:
        try:
            rows = con.execute(f'SELECT * FROM "{name}" ORDER BY rowid DESC LIMIT 3').fetchall()
            sample = [list(row) for row in rows]
        except sqlite3.DatabaseError:
            sample = []
    return {"name": name, "columns": cols, "row_count": count, "sample_tail": sample}


def inspect_database(path: Path) -> dict[str, Any]:
    con = sqlite3.connect(path)
    try:
        tables = [row[0] for row in con.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")]
        integrity = con.execute("PRAGMA integrity_check").fetchone()[0]
        fk = list(con.execute("PRAGMA foreign_key_check"))
        relevant = [
            name for name in tables
            if any(token in name.lower() for token in (
                "response", "prompt", "checkpoint", "recovery", "release", "tracking", "instruction", "print", "cover", "spine"
            ))
        ]
        states: dict[str, Any] = {}
        for name in ("section4_final_release", "section4_session3_checkpoint3_acceptance", "thread_response_reconciliation_cp3", "fractional_prompt_cp3"):
            if name in tables:
                states[name] = table_schema(con, name)
        return {
            "path": str(path),
            "relative_path": None,
            "bytes": path.stat().st_size,
            "sha256": sha256_file(path),
            "table_count": len(tables),
            "integrity": integrity,
            "foreign_key_violations": len(fk),
            "relevant_tables": relevant,
            "selected_schemas": states,
        }
    finally:
        con.close()


def workbook_sheets(path: Path) -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--volume1-dir", type=Path, required=True)
    parser.add_argument("--volume2-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, default=Path("inspection_output"))
    args = parser.parse_args()
    out = args.output_dir
    out.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory(prefix="mrhpd-r72-inspect-") as td:
        work = Path(td)
        restore = reconstruct_restore(args.volume1_dir, args.volume2_dir, work)
        restore_root = work / "restore"
        safe_extract(restore, restore_root)
        project_archive = locate_project_archive(restore_root)
        project_extract = work / "project_extract"
        safe_extract(project_archive, project_extract)
        roots = [p for p in project_extract.iterdir() if p.is_dir()]
        project = roots[0] if len(roots) == 1 else project_extract

        db_candidates = sorted(
            [p for p in project.rglob("*") if p.is_file() and p.suffix.lower() in {".sqlite", ".db", ".sqlite3"}],
            key=lambda p: p.stat().st_size,
            reverse=True,
        )
        db_reports = []
        for path in db_candidates:
            try:
                report = inspect_database(path)
                report["relative_path"] = path.relative_to(project).as_posix()
                db_reports.append(report)
            except Exception as exc:
                db_reports.append({"relative_path": path.relative_to(project).as_posix(), "error": repr(exc), "bytes": path.stat().st_size})

        xlsx_candidates = sorted(project.rglob("*.xlsx"), key=lambda p: p.stat().st_size, reverse=True)
        workbook_reports = []
        for path in xlsx_candidates:
            try:
                sheets = workbook_sheets(path)
                workbook_reports.append({
                    "relative_path": path.relative_to(project).as_posix(),
                    "bytes": path.stat().st_size,
                    "sha256": sha256_file(path),
                    "sheet_count": len(sheets),
                    "sheets": sheets,
                })
            except Exception as exc:
                workbook_reports.append({"relative_path": path.relative_to(project).as_posix(), "error": repr(exc)})

        pdfs = []
        for path in sorted(project.rglob("*.pdf"), key=lambda p: p.stat().st_size, reverse=True):
            pdfs.append({"relative_path": path.relative_to(project).as_posix(), "bytes": path.stat().st_size, "sha256": sha256_file(path)})

        covers = []
        for path in sorted(project.rglob("*")):
            if path.is_file() and any(token in path.name.lower() for token in ("cover", "spine", "wrap")):
                covers.append({"relative_path": path.relative_to(project).as_posix(), "bytes": path.stat().st_size, "sha256": sha256_file(path)})

        tracking = []
        for path in sorted(project.rglob("*")):
            if path.is_file() and any(token in path.as_posix().lower() for token in ("tracking", "thread index", "raw prompt", "net prompt", "raw response", "net response")):
                tracking.append({"relative_path": path.relative_to(project).as_posix(), "bytes": path.stat().st_size})

        instructions = []
        for path in sorted(project.rglob("Instructions.*")):
            if path.is_file():
                instructions.append({"relative_path": path.relative_to(project).as_posix(), "bytes": path.stat().st_size, "sha256": sha256_file(path)})

        files = [p for p in project.rglob("*") if p.is_file()]
        report = {
            "schema": "mrhpd-section5-intake-inspection-1.0",
            "status": "passed",
            "restore": {"name": restore.name, "bytes": restore.stat().st_size, "sha256": sha256_file(restore)},
            "project_archive": {"name": project_archive.name, "bytes": project_archive.stat().st_size, "sha256": sha256_file(project_archive)},
            "project_root_name": project.name,
            "physical_file_count": len(files),
            "physical_bytes": sum(p.stat().st_size for p in files),
            "databases": db_reports,
            "workbooks": workbook_reports,
            "pdfs": pdfs[:30],
            "cover_spine_wrap_files": covers,
            "tracking_files": tracking,
            "instruction_files": instructions,
        }
        (out / "MRHPD_RESPONSE72_SECTION5_INTAKE_INSPECTION.json").write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
        lines = [
            "MRHPD Response 72 Section 5 Intake Inspection",
            f"Status: {report['status']}",
            f"Project files: {report['physical_file_count']}",
            f"Database candidates: {len(db_reports)}",
            f"Workbook candidates: {len(workbook_reports)}",
            f"PDF candidates: {len(pdfs)}",
            f"Cover/spine/wrap files: {len(covers)}",
            "",
            "DATABASES",
        ]
        for row in db_reports:
            lines.append(json.dumps({k: row.get(k) for k in ("relative_path", "bytes", "sha256", "table_count", "integrity", "foreign_key_violations", "error")}, ensure_ascii=False))
        lines.append("\nWORKBOOKS")
        for row in workbook_reports:
            lines.append(json.dumps({k: row.get(k) for k in ("relative_path", "bytes", "sha256", "sheet_count", "error")}, ensure_ascii=False))
        lines.append("\nCOVER/SPINE/WRAP")
        lines.extend(row["relative_path"] for row in covers)
        (out / "MRHPD_RESPONSE72_SECTION5_INTAKE_INSPECTION.txt").write_text("\n".join(lines), encoding="utf-8")
        print(json.dumps({"status": "passed", "output": str(out), "files": [p.name for p in out.iterdir()]}, indent=2))


if __name__ == "__main__":
    main()
