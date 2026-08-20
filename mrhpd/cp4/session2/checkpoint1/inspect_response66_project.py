#!/usr/bin/env python3
"""Inspect the exact MRHPD complete restore through Response 66.

This utility is read-only. It reconstructs the two-volume complete restore,
verifies the governed identities, extracts a disposable copy, and emits a
machine-readable map for Section 4 Session 2 Checkpoint 1 planning.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
import subprocess
import sys
import tempfile
import zipfile
from collections import Counter
from pathlib import Path, PurePosixPath
from typing import Any

RESPONSE66_RESTORE_BYTES = 177_617_796
RESPONSE66_RESTORE_SHA256 = "38c8fa08763d5698217ce33a2bbe1e889e726087575b14fb31086f38cfe1300f"
PROJECT_SNAPSHOT_BYTES = 169_294_854
PROJECT_SNAPSHOT_SHA256 = "b59e5265c0515a5dbaadf55b631a37c581b828b1a37857ee3322cda532125cc4"
PUBLICATION_SHA256 = "8a053112ca24cd730b970130d5d0fc57a15c681531603601096186aeb0cd9642"
EDITABLE_ASSEMBLY_SHA256 = "f832ff934d77049d75712f28bdfc9167b8a6b119c797235431b304b9e24369a2"


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(chunk_size), b""):
            h.update(block)
    return h.hexdigest()


def safe_infos(zf: zipfile.ZipFile) -> list[zipfile.ZipInfo]:
    infos = zf.infolist()
    names = [i.filename for i in infos]
    duplicates = [name for name, count in Counter(names).items() if count > 1]
    unsafe: list[str] = []
    for name in names:
        p = PurePosixPath(name.replace("\\", "/"))
        if name.startswith(("/", "\\")) or re.match(r"^[A-Za-z]:", name) or ".." in p.parts:
            unsafe.append(name)
    if duplicates or unsafe:
        raise RuntimeError({"duplicates": duplicates[:20], "unsafe": unsafe[:20]})
    bad = zf.testzip()
    if bad:
        raise RuntimeError({"zip_crc_error": bad})
    return infos


def verify_zip(path: Path, expected_bytes: int | None = None, expected_sha256: str | None = None) -> dict[str, Any]:
    if expected_bytes is not None and path.stat().st_size != expected_bytes:
        raise RuntimeError({"file": str(path), "expected_bytes": expected_bytes, "actual_bytes": path.stat().st_size})
    digest = sha256_file(path)
    if expected_sha256 is not None and digest != expected_sha256:
        raise RuntimeError({"file": str(path), "expected_sha256": expected_sha256, "actual_sha256": digest})
    with zipfile.ZipFile(path) as zf:
        infos = safe_infos(zf)
    return {"file": path.name, "bytes": path.stat().st_size, "sha256": digest, "members": len(infos), "crc": "passed"}


def extract(path: Path, target: Path) -> None:
    target.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path) as zf:
        safe_infos(zf)
        zf.extractall(target)


def find_one(root: Path, pattern: str) -> Path:
    matches = sorted(root.rglob(pattern))
    if len(matches) != 1:
        raise RuntimeError({"pattern": pattern, "matches": [str(p) for p in matches]})
    return matches[0]


def reconstruct_restore(input_root: Path, work: Path) -> tuple[Path, dict[str, Any]]:
    wrappers = [
        find_one(input_root, "*Response 66 Complete Restore Drive Volume 1 of 2.zip"),
        find_one(input_root, "*Response 66 Complete Restore Drive Volume 2 of 2.zip"),
    ]
    wrapper_qa = [verify_zip(path) for path in wrappers]
    extracted: list[Path] = []
    for index, wrapper in enumerate(wrappers, 1):
        target = work / f"volume_{index}"
        extract(wrapper, target)
        extracted.append(target)
    staging = work / "reassembly"
    staging.mkdir()
    common_names: dict[str, str] = {}
    for target in extracted:
        for path in target.rglob("*"):
            if not path.is_file():
                continue
            destination = staging / path.name
            digest = sha256_file(path)
            if destination.exists():
                if common_names[path.name] != digest:
                    raise RuntimeError({"duplicate_common_file_mismatch": path.name})
                continue
            destination.write_bytes(path.read_bytes())
            common_names[path.name] = digest
    utility = find_one(staging, "reassemble_response66_complete_restore.py")
    result = subprocess.run([sys.executable, str(utility)], cwd=staging, text=True, capture_output=True, timeout=1200)
    if result.returncode != 0:
        raise RuntimeError({"reassembly_failed": {"stdout": result.stdout[-10000:], "stderr": result.stderr[-10000:]}})
    restores = [p for p in staging.glob("*.zip") if p.stat().st_size == RESPONSE66_RESTORE_BYTES]
    if len(restores) != 1:
        raise RuntimeError({"restore_candidates": [(p.name, p.stat().st_size) for p in staging.glob("*.zip")]})
    restore = restores[0]
    restore_qa = verify_zip(restore, RESPONSE66_RESTORE_BYTES, RESPONSE66_RESTORE_SHA256)
    return restore, {"wrappers": wrapper_qa, "reassembly_stdout": result.stdout[-5000:], "restore": restore_qa}


def locate_project_snapshot(restore: Path, work: Path) -> tuple[Path, Path, dict[str, Any]]:
    restore_root = work / "restore"
    extract(restore, restore_root)
    candidates = sorted(restore_root.rglob("*.zip"), key=lambda p: p.stat().st_size, reverse=True)
    matches = [p for p in candidates if p.stat().st_size == PROJECT_SNAPSHOT_BYTES and sha256_file(p) == PROJECT_SNAPSHOT_SHA256]
    if len(matches) != 1:
        raise RuntimeError({"project_snapshot_matches": [(str(p), p.stat().st_size) for p in matches]})
    snapshot = matches[0]
    snapshot_qa = verify_zip(snapshot, PROJECT_SNAPSHOT_BYTES, PROJECT_SNAPSHOT_SHA256)
    project_extract = work / "project"
    extract(snapshot, project_extract)
    roots = [p for p in project_extract.iterdir() if p.is_dir()]
    files = [p for p in project_extract.iterdir() if p.is_file()]
    project = roots[0] if len(roots) == 1 and not files else project_extract
    return snapshot, project, snapshot_qa


def table_exists(con: sqlite3.Connection, table: str) -> bool:
    return con.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone() is not None


def find_database(project: Path) -> Path:
    candidates = sorted(project.rglob("*.sqlite"), key=lambda p: p.stat().st_size, reverse=True)
    scored: list[tuple[int, int, Path]] = []
    for path in candidates:
        try:
            con = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
            tables = con.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table'").fetchone()[0]
            max_response = 0
            if table_exists(con, "thread_response_reconciliation_cp3"):
                max_response = con.execute("SELECT COALESCE(MAX(response_number),0) FROM thread_response_reconciliation_cp3").fetchone()[0]
            con.close()
            scored.append((int(max_response or 0), int(tables or 0), path))
        except Exception:
            continue
    if not scored:
        raise FileNotFoundError("No readable SQLite database")
    scored.sort(key=lambda row: (row[0], row[1], row[2].stat().st_size), reverse=True)
    return scored[0][2]


def schema(con: sqlite3.Connection, table: str) -> list[dict[str, Any]]:
    if not table_exists(con, table):
        return []
    return [
        {"cid": row[0], "name": row[1], "type": row[2], "notnull": row[3], "default": row[4], "pk": row[5]}
        for row in con.execute(f'PRAGMA table_info("{table}")')
    ]


def safe_count(con: sqlite3.Connection, table: str) -> int | None:
    try:
        return int(con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0])
    except Exception:
        return None


def inspect_database(db: Path, project: Path) -> dict[str, Any]:
    con = sqlite3.connect(f"file:{db}?mode=ro", uri=True)
    try:
        integrity = con.execute("PRAGMA integrity_check").fetchone()[0]
        fk = list(con.execute("PRAGMA foreign_key_check"))
        tables = [row[0] for row in con.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")]
        counts = {name: safe_count(con, name) for name in tables}
        selected = [
            "thread_response_reconciliation_cp3", "fractional_prompt_cp3", "remediation_recovery_event",
            "section4_checkpoint", "section4_session_release", "section4_sync_qa", "restore_emission_policy",
            "metadata", "evidence_source", "taxonomy_node", "search_resolver_rule", "search_document",
            "organism_alias", "organism_clinical_profile", "morphology_profile", "lab_growth_profile",
            "transmission_epidemiology_profile", "organism_common_source", "organism_disease_association",
            "manifestation", "diagnostic_test", "organism_diagnostic_test", "organism_resistance_profile",
            "treatment_context", "treatment_option", "treatment_option_detail", "treatment_duration_rule",
            "stewardship_rule", "graphic_asset", "graphic_prompt", "publication_page_map",
            "publication_index_locator", "publication_cross_reference", "master_category",
        ]
        schemas = {name: schema(con, name) for name in selected if table_exists(con, name)}
        response_max = counts.get("thread_response_reconciliation_cp3")
        current_response = None
        if table_exists(con, "thread_response_reconciliation_cp3"):
            current_response = con.execute("SELECT MAX(response_number) FROM thread_response_reconciliation_cp3").fetchone()[0]
        metadata_rows = []
        if table_exists(con, "metadata"):
            cols = [row[1] for row in con.execute('PRAGMA table_info("metadata")')]
            if {"key", "value"}.issubset(cols):
                metadata_rows = [dict(zip(["key", "value"], row)) for row in con.execute("SELECT key,value FROM metadata ORDER BY key")]
        return {
            "path": db.relative_to(project).as_posix(),
            "bytes": db.stat().st_size,
            "sha256": sha256_file(db),
            "integrity": integrity,
            "foreign_key_violations": len(fk),
            "table_count": len(tables),
            "tables": tables,
            "counts": counts,
            "selected_schemas": schemas,
            "current_response": current_response,
            "metadata": metadata_rows,
        }
    finally:
        con.close()


def inspect_workbooks(project: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook
    rows: list[dict[str, Any]] = []
    for path in sorted(project.rglob("*.xlsx"), key=lambda p: p.stat().st_size, reverse=True)[:20]:
        try:
            wb = load_workbook(path, read_only=True, data_only=False)
            sheets = list(wb.sheetnames)
            formula_count = 0
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            formula_count += 1
            wb.close()
            rows.append({
                "path": path.relative_to(project).as_posix(),
                "bytes": path.stat().st_size,
                "sha256": sha256_file(path),
                "sheet_count": len(sheets),
                "sheets": sheets,
                "formula_count": formula_count,
            })
        except Exception as exc:
            rows.append({"path": path.relative_to(project).as_posix(), "error": repr(exc)})
    return rows


def inspect_application(project: Path, db: Path) -> dict[str, Any]:
    apps = sorted(project.rglob("human_pathogen_app.py"))
    app = apps[0] if len(apps) == 1 else None
    if app is None:
        return {"matches": [str(p.relative_to(project)) for p in apps]}
    source = app.read_text(encoding="utf-8", errors="replace")
    endpoint_tokens = sorted(set(re.findall(r"['\"]([a-z][a-z0-9_/-]{2,60})['\"]", source)))
    endpoint_tokens = [token for token in endpoint_tokens if any(word in token for word in ("search", "organism", "syndrome", "treatment", "pathway", "evidence", "graphic", "status", "summary", "cross", "publication", "resolver", "taxonomy", "health"))]
    help_result = subprocess.run([sys.executable, str(app), "--help"], cwd=app.parent, text=True, capture_output=True, timeout=120)
    tests = [p.relative_to(project).as_posix() for p in sorted(app.parent.glob("test*.py"))]
    state_files = [p.relative_to(project).as_posix() for p in sorted(app.parent.glob("CURRENT*"))]
    return {
        "path": app.relative_to(project).as_posix(),
        "bytes": app.stat().st_size,
        "sha256": sha256_file(app),
        "help_returncode": help_result.returncode,
        "help_stdout": help_result.stdout[-12000:],
        "help_stderr": help_result.stderr[-4000:],
        "candidate_endpoint_tokens": endpoint_tokens[:250],
        "tests": tests,
        "state_files": state_files,
        "database": db.relative_to(project).as_posix(),
    }


def inspect_publication(project: Path) -> dict[str, Any]:
    pdfs = [p for p in project.rglob("*Integrated Manuscript*.pdf") if sha256_file(p) == PUBLICATION_SHA256]
    docxs = [p for p in project.rglob("*Editable Integrated Manuscript Assembly*.docx") if sha256_file(p) == EDITABLE_ASSEMBLY_SHA256]
    return {
        "publication_matches": [p.relative_to(project).as_posix() for p in pdfs],
        "editable_assembly_matches": [p.relative_to(project).as_posix() for p in docxs],
        "publication_sha256": PUBLICATION_SHA256,
        "editable_assembly_sha256": EDITABLE_ASSEMBLY_SHA256,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", type=Path, default=Path("response66_artifacts"))
    parser.add_argument("--output-dir", type=Path, default=Path("inspection_output"))
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="mrhpd-s4s2-inspect-") as td:
        work = Path(td)
        restore, restore_qa = reconstruct_restore(args.input_dir, work)
        snapshot, project, snapshot_qa = locate_project_snapshot(restore, work)
        db = find_database(project)
        payload = {
            "schema": "mrhpd-section4-session2-response66-inspection-1.0",
            "status": "passed",
            "read_only": True,
            "accepted_predecessor_mutated": False,
            "restore": restore_qa,
            "project_snapshot": snapshot_qa,
            "project_root": project.name,
            "project_file_count": sum(1 for p in project.rglob("*") if p.is_file()),
            "project_total_bytes": sum(p.stat().st_size for p in project.rglob("*") if p.is_file()),
            "database": inspect_database(db, project),
            "workbooks": inspect_workbooks(project),
            "application": inspect_application(project, db),
            "publication": inspect_publication(project),
            "tracking_files": [p.relative_to(project).as_posix() for p in sorted(project.rglob("*Tracking*")) if p.is_file()][:500],
            "qa_files": [p.relative_to(project).as_posix() for p in sorted(project.rglob("QA/**/*")) if p.is_file()][:1000],
            "recovery_files": [p.relative_to(project).as_posix() for p in sorted(project.rglob("Recovery/**/*")) if p.is_file()][:1000],
            "top_level": sorted(p.name for p in project.iterdir()),
        }
        output = args.output_dir / "MRHPD_RESPONSE66_PROJECT_INSPECTION.json"
        output.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        print(json.dumps({
            "status": payload["status"],
            "output": str(output),
            "database": payload["database"]["path"],
            "tables": payload["database"]["table_count"],
            "current_response": payload["database"]["current_response"],
            "workbook_count": len(payload["workbooks"]),
            "largest_workbook_sheets": payload["workbooks"][0].get("sheet_count") if payload["workbooks"] else None,
            "application": payload["application"].get("path"),
        }, indent=2))


if __name__ == "__main__":
    main()
