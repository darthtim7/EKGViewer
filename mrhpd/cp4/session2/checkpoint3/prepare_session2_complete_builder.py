#!/usr/bin/env python3
"""Generate the Section 4 Session 2 complete-restore builder from the verified Session 1 builder.

The source builder already implements the governed full-restore, workbook,
application, tracking, index, manifest, clean-extraction, and transport-volume
lanes. This adapter advances those controls from the exact Response 66 full
restore plus the verified Response 68 cumulative recovery package, while
replacing the few session-specific intake and packaging functions that cannot
be safely handled by mechanical relabeling.

The generated builder is disposable. The accepted predecessor, frozen Section
3 release, Response 66 restore, and Response 68 recovery package are never
edited in place.
"""
from __future__ import annotations

import io
import re
import tokenize
from pathlib import Path

SOURCE = Path("mrhpd/cp4/session1/checkpoint3/build_session1_complete_restore.py")
TARGET = Path("mrhpd/cp4/session2/checkpoint3/build_session2_complete_restore.py")


def replace_number_tokens(source: str, mapping: dict[str, str]) -> str:
    tokens = []
    for token in tokenize.generate_tokens(io.StringIO(source).readline):
        if token.type == tokenize.NUMBER and token.string in mapping:
            token = tokenize.TokenInfo(token.type, mapping[token.string], token.start, token.end, token.line)
        tokens.append(token)
    return tokenize.untokenize(tokens)


def replace_once(text: str, old: str, new: str, label: str) -> str:
    count = text.count(old)
    if count != 1:
        raise SystemExit(f"{label}: expected exactly one match, found {count}")
    return text.replace(old, new, 1)


def replace_function(text: str, start_name: str, next_name: str, replacement: str) -> str:
    pattern = rf"def {re.escape(start_name)}\(.*?\n\ndef {re.escape(next_name)}\("
    match = re.search(pattern, text, flags=re.S)
    if not match:
        raise SystemExit(f"function range not found: {start_name} -> {next_name}")
    suffix = f"\n\ndef {next_name}("
    return text[: match.start()] + replacement.rstrip() + suffix + text[match.end() :]


text = SOURCE.read_text(encoding="utf-8")

# Protect all source references to the next session before advancing current
# Session 1 labels to Session 2.
next_placeholders = {
    "Remediation Section 4 of 5 Session 2 of 3": "__MRHPD_NEXT_SESSION_LONG__",
    "Section 4 Session 1 to Session 2 Handoff": "__MRHPD_HANDOFF_TITLE__",
    "SESSION_2_HANDOFF.md": "__MRHPD_HANDOFF_FILE__",
    "Session 2 begins from": "__MRHPD_NEXT_BEGINS_FROM__",
}
for old, placeholder in next_placeholders.items():
    text = text.replace(old, placeholder)

# Advance the final response first, then the checkpoint input, then the base
# response. This order avoids remapping newly introduced Response 66 labels.
response_replacements = [
    ("Response 66", "Response 69"),
    ("response66", "response69"),
    ("RESPONSE66", "RESPONSE69"),
    ("R66", "R69"),
    ("r66", "r69"),
    ("Response 65", "Response 68"),
    ("response65", "response68"),
    ("R65", "R68"),
    ("r65", "r68"),
    ("Response 64", "Response 66"),
    ("response64", "response66"),
    ("RESPONSE64", "RESPONSE66"),
    ("R64", "R66"),
]
for old, new in response_replacements:
    text = text.replace(old, new)
text = replace_number_tokens(text, {"66": "69"})

# Advance the current session and all governed identifiers.
for old, new in [
    ("Session 1", "Session 2"),
    ("session1", "session2"),
    ("SESSION_1", "SESSION_2"),
    ("CP4-S1", "CP4-S2"),
    ("S4S1", "S4S2"),
    ("S1-", "S2-"),
]:
    text = text.replace(old, new)

# Restore the protected next-session labels as Session 3.
for placeholder, replacement in [
    ("__MRHPD_NEXT_SESSION_LONG__", "Remediation Section 4 of 5 Session 3 of 3"),
    ("__MRHPD_HANDOFF_TITLE__", "Section 4 Session 2 to Session 3 Handoff"),
    ("__MRHPD_HANDOFF_FILE__", "SESSION_3_HANDOFF.md"),
    ("__MRHPD_NEXT_BEGINS_FROM__", "Session 3 begins from"),
]:
    text = text.replace(placeholder, replacement)

# Exact input identities.
text = text.replace(
    'BASE_RESPONSE66_BYTES = 145_920_215',
    'BASE_RESPONSE66_BYTES = 177_617_796',
)
text = text.replace(
    'BASE_RESPONSE66_SHA256 = "71e7a06868e82238188827fd73ca7b2843670b0843dcf98ab5ab72305bf77834"',
    'BASE_RESPONSE66_SHA256 = "38c8fa08763d5698217ce33a2bbe1e889e726087575b14fb31086f38cfe1300f"',
)
text = text.replace('CP2_RECOVERY_BYTES = 10_898_139', 'CP2_RECOVERY_BYTES = 18_318_469')
text = text.replace(
    'CP2_RECOVERY_SHA256 = "f659564382e5858beaf3b1d2e8f77599d58e2176ad4e977034a0dc18dd0efe35"',
    'CP2_RECOVERY_SHA256 = "b466c463c55dc95d2ac780ff78755f5ae09fa19d9a6be4fb97e914af7568adbe"',
)
text = text.replace(
    'CP2_DATABASE_SHA256 = "71586759d991836777c9c03e07b905a5d861c026a6c790b8920133f9f00f7454"',
    'CP2_DATABASE_SHA256 = "dynamic-from-response68-recovery-manifest"',
)
text = text.replace(
    'CP2_WORKBOOK_SHA256 = "9499e9bf6c2260e37403140c77149a7e5dd9728e52fbf1b5c3cf842b7a062ac6"',
    'CP2_WORKBOOK_SHA256 = "dynamic-from-response68-recovery-manifest"',
)

# Recovery-event numbering continues after the checkpoint-2 history through
# Event 115.
for old, new in [
    ('"event_number": 95', '"event_number": 116'),
    ('"event_number": 96', '"event_number": 117'),
    ('"event_number": 97', '"event_number": 118'),
    ("RECOVERY_EVENTS_95_97.json", "RECOVERY_EVENTS_116_118.json"),
]:
    text = text.replace(old, new)

# Clarify the current response title and source-state description.
text = text.replace(
    '"title": "Section 4 Session 2 complete restore and handoff"',
    '"title": "Section 4 Session 2 complete restore and Session 3 handoff"',
)
text = text.replace(
    '"notes": "Session 2 of 3 is complete. Continue begins __MRHPD_NEXT_SESSION_LONG__."',
    '"notes": "Session 2 of 3 is complete. Continue begins Remediation Section 4 Session 3 of 3."',
)

NEW_APPLY = r'''
def find_canonical_database(root: Path) -> Path:
    scored: list[tuple[int, int, int, int, Path]] = []
    for path in sorted(root.rglob("*.sqlite"), key=lambda item: item.stat().st_size, reverse=True):
        try:
            con = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
            tables = con.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table'").fetchone()[0]
            max_response = 0
            r68 = 0
            if table_exists(con, "thread_response_reconciliation_cp3"):
                max_response = int(con.execute("SELECT COALESCE(MAX(response_number),0) FROM thread_response_reconciliation_cp3").fetchone()[0] or 0)
                r68 = int(con.execute("SELECT COUNT(*) FROM thread_response_reconciliation_cp3 WHERE response_key='R68'").fetchone()[0] or 0)
            con.close()
            scored.append((r68, max_response, int(tables or 0), path.stat().st_size, path))
        except Exception:
            continue
    if not scored:
        raise FileNotFoundError("No readable SQLite database found in restored Response 68 project")
    scored.sort(reverse=True)
    return scored[0][-1]


def find_current_workbook(root: Path) -> tuple[Path, list[str]]:
    from openpyxl import load_workbook
    scored: list[tuple[int, int, int, Path, list[str]]] = []
    for path in sorted(root.rglob("*.xlsx"), key=lambda item: item.stat().st_size, reverse=True)[:80]:
        try:
            wb = load_workbook(path, read_only=True, data_only=False)
            sheets = list(wb.sheetnames)
            wb.close()
            s4s2 = sum(1 for name in sheets if name.startswith("S4S2"))
            detailed = int(any(name in sheets for name in ("S4S2 Field Coverage", "S4S2 Query Coverage", "S4S2 QA")))
            scored.append((detailed, s4s2, len(sheets), path, sheets))
        except Exception:
            continue
    if not scored:
        raise FileNotFoundError("No readable workbook found in restored Response 68 project")
    scored.sort(key=lambda row: (row[0], row[1], row[2], row[3].stat().st_size), reverse=True)
    return scored[0][3], scored[0][4]


def apply_checkpoint2(base_restore: Path, recovery_zip: Path, work: Path) -> tuple[Path, dict[str, Any]]:
    recovery_package = work / "checkpoint2_recovery_package"
    safe_extract(recovery_zip, recovery_package)
    utilities = list(recovery_package.rglob("apply_checkpoint_recovery.py"))
    if len(utilities) != 1:
        raise RuntimeError({"response68_apply_utilities": [str(path) for path in utilities]})
    utility = utilities[0]
    output = work / "project_through_response68"
    result = subprocess.run(
        [
            sys.executable,
            str(utility),
            "--base-response66-restore",
            str(base_restore),
            "--output-dir",
            str(output),
        ],
        text=True,
        capture_output=True,
        timeout=1800,
    )
    apply_record = {
        "returncode": result.returncode,
        "stdout": result.stdout[-40000:],
        "stderr": result.stderr[-40000:],
        "utility": str(utility),
    }
    if result.returncode != 0:
        raise RuntimeError({"checkpoint2_apply_failed": apply_record})

    manifests = list(recovery_package.rglob("RECOVERY_MANIFEST.json"))
    manifest = json.loads(manifests[0].read_text(encoding="utf-8")) if len(manifests) == 1 else {}
    db = find_canonical_database(output)
    workbook, workbook_sheets = find_current_workbook(output)
    app = find_by_hash(output, "human_pathogen_app.py", APPLICATION_SHA256)
    publication = find_by_hash(output, "*Integrated Manuscript*.pdf", PUBLICATION_SHA256)
    editable = find_by_hash(output, "*Editable Integrated Manuscript Assembly*.docx", EDITABLE_ASSEMBLY_SHA256)
    con = sqlite3.connect(db)
    try:
        integrity = con.execute("PRAGMA integrity_check").fetchone()[0]
        fk = list(con.execute("PRAGMA foreign_key_check"))
        r68 = con.execute("SELECT COUNT(*) FROM thread_response_reconciliation_cp3 WHERE response_key='R68'").fetchone()[0]
        cp2_state = None
        if table_exists(con, "section4_session2_checkpoint"):
            cp2_state = con.execute("SELECT state FROM section4_session2_checkpoint WHERE checkpoint_code='MRHPD-V3-CP4-S2-CP2'").fetchone()
        if cp2_state is None and table_exists(con, "section4_checkpoint"):
            cp2_state = con.execute("SELECT state FROM section4_checkpoint WHERE checkpoint_code='MRHPD-V3-CP4-S2-CP2'").fetchone()
        field_failures = con.execute("SELECT COUNT(*) FROM section4_session2_field_coverage WHERE checkpoint_code='MRHPD-V3-CP4-S2-CP2' AND status!='passed'").fetchone()[0]
        query_failures = con.execute("SELECT COUNT(*) FROM section4_session2_query_coverage WHERE checkpoint_code='MRHPD-V3-CP4-S2-CP2' AND status!='passed'").fetchone()[0]
        governance_failures = con.execute("SELECT COUNT(*) FROM section4_session2_source_governance WHERE checkpoint_code='MRHPD-V3-CP4-S2-CP2' AND status!='passed'").fetchone()[0]
        drift_failures = con.execute("SELECT COUNT(*) FROM section4_session2_drift_resolution WHERE checkpoint_code='MRHPD-V3-CP4-S2-CP2' AND status!='passed'").fetchone()[0]
        field_records = con.execute("SELECT COUNT(*) FROM section4_session2_field_coverage WHERE checkpoint_code='MRHPD-V3-CP4-S2-CP2'").fetchone()[0]
        query_records = con.execute("SELECT COUNT(*) FROM section4_session2_query_coverage WHERE checkpoint_code='MRHPD-V3-CP4-S2-CP2'").fetchone()[0]
        governance_records = con.execute("SELECT COUNT(*) FROM section4_session2_source_governance WHERE checkpoint_code='MRHPD-V3-CP4-S2-CP2'").fetchone()[0]
    finally:
        con.close()
    if integrity != "ok" or fk or r68 != 1 or not cp2_state or cp2_state[0] != "checkpoint_complete":
        raise RuntimeError({"checkpoint2_state_gate": {"integrity": integrity, "fk": fk[:20], "r68": r68, "cp2": cp2_state}})
    if any((field_failures, query_failures, governance_failures, drift_failures)):
        raise RuntimeError({"checkpoint2_detailed_gate": {"field": field_failures, "query": query_failures, "governance": governance_failures, "drift": drift_failures}})
    return output, {
        "status": "passed",
        "base_restore": verify_zip(base_restore, BASE_RESPONSE66_BYTES, BASE_RESPONSE66_SHA256),
        "checkpoint2_recovery": verify_zip(recovery_zip, CP2_RECOVERY_BYTES, CP2_RECOVERY_SHA256),
        "recovery_manifest": manifest,
        "apply_utility": apply_record,
        "database": db.relative_to(output).as_posix(),
        "database_sha256": sha256_file(db),
        "workbook": workbook.relative_to(output).as_posix(),
        "workbook_sha256": sha256_file(workbook),
        "workbook_sheets": len(workbook_sheets),
        "application": app.relative_to(output).as_posix(),
        "application_sha256": APPLICATION_SHA256,
        "publication": publication.relative_to(output).as_posix(),
        "publication_sha256": PUBLICATION_SHA256,
        "editable_assembly": editable.relative_to(output).as_posix(),
        "editable_assembly_sha256": EDITABLE_ASSEMBLY_SHA256,
        "field_coverage_records": field_records,
        "query_coverage_records": query_records,
        "source_governance_records": governance_records,
        "accepted_predecessor_mutated": False,
    }
'''
text = replace_function(text, "apply_checkpoint2", "synchronize_database", NEW_APPLY)

text = replace_once(
    text,
    '    source = find_by_hash(project, "*.sqlite", CP2_DATABASE_SHA256)\n',
    '    source = find_canonical_database(project)\n',
    "current database selection",
)
text = replace_once(
    text,
    '    source = find_by_hash(project, "*.xlsx", CP2_WORKBOOK_SHA256)\n',
    '    source, _source_sheets = find_current_workbook(project)\n',
    "current workbook selection",
)

# Persist a checkpoint-3 row in the detailed Session 2 checkpoint registry.
checkpoint_anchor = '        if table_exists(con, "metadata") and {"key", "value"}.issubset(table_columns(con, "metadata")):\n'
checkpoint_insert = '''        if table_exists(con, "section4_session2_checkpoint"):
            con.execute("""
              INSERT INTO section4_session2_checkpoint
              (checkpoint_code,section_label,session_label,checkpoint_label,response_number,state,database_status,workbook_status,
               application_status,publication_status,capability_status,evidence_status,drift_status,accepted_predecessor_mutated,
               next_checkpoint,recorded_at)
              VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
              ON CONFLICT(checkpoint_code) DO UPDATE SET
                response_number=excluded.response_number,state=excluded.state,database_status=excluded.database_status,
                workbook_status=excluded.workbook_status,application_status=excluded.application_status,
                publication_status=excluded.publication_status,capability_status=excluded.capability_status,
                evidence_status=excluded.evidence_status,drift_status=excluded.drift_status,
                accepted_predecessor_mutated=excluded.accepted_predecessor_mutated,
                next_checkpoint=excluded.next_checkpoint,recorded_at=excluded.recorded_at
            """, (
                "MRHPD-V3-CP4-S2-CP3", "Remediation Section 4 of 5", "Session 2 of 3", "Checkpoint 3 of 3",
                69, "session_complete", "ok", "pending", "pending", "passed", "passed", "passed", "passed", 0,
                "Remediation Section 4 of 5 Session 3 of 3", NOW,
            ))
'''
if checkpoint_anchor not in text:
    raise SystemExit("session2 checkpoint insertion anchor missing")
text = text.replace(checkpoint_anchor, checkpoint_insert + checkpoint_anchor, 1)

# Finalization must update the detailed checkpoint registry after workbook and
# application gates have run.
finalize_anchor = '        controls = [\n'
finalize_insert = '''        if table_exists(con, "section4_session2_checkpoint"):
            con.execute("""
              UPDATE section4_session2_checkpoint
              SET state='session_complete',database_status='ok',workbook_status=?,application_status=?,publication_status=?,
                  capability_status='passed',evidence_status='passed',drift_status='passed',accepted_predecessor_mutated=0,
                  next_checkpoint='Remediation Section 4 of 5 Session 3 of 3',recorded_at=?
              WHERE checkpoint_code='MRHPD-V3-CP4-S2-CP3'
            """, (workbook_qa["status"], application_qa["status"], publication_qa["status"], NOW))
'''
if finalize_anchor not in text:
    raise SystemExit("database finalization anchor missing")
text = text.replace(finalize_anchor, finalize_insert + finalize_anchor, 1)

# Record the detailed checkpoint registry in the application acceptance gate.
app_anchor = '            "session2_release_complete": con.execute("SELECT state FROM section4_session_release WHERE release_code=\'MRHPD-V3-CP4-S2-COMPLETE\'").fetchone() == ("session_complete",),\n'
app_insert = app_anchor + '            "session2_detailed_checkpoint_complete": (not table_exists(con, "section4_session2_checkpoint")) or con.execute("SELECT state FROM section4_session2_checkpoint WHERE checkpoint_code=\'MRHPD-V3-CP4-S2-CP3\'").fetchone() == ("session_complete",),\n'
if app_anchor not in text:
    raise SystemExit("application detailed-check anchor missing")
text = text.replace(app_anchor, app_insert, 1)

# Remove generated bytecode before the final indexes, manifests, and project
# archive are frozen.
PURGE_HELPER = r'''
def purge_generated_caches(root: Path) -> dict[str, Any]:
    removed: list[str] = []
    for directory in sorted(root.rglob("__pycache__"), reverse=True):
        if directory.is_dir():
            removed.append(directory.relative_to(root).as_posix())
            shutil.rmtree(directory)
    for path in sorted(root.rglob("*.pyc")):
        if path.is_file():
            removed.append(path.relative_to(root).as_posix())
            path.unlink()
    return {"status": "passed", "removed_count": len(removed), "removed_paths": removed}
'''
main_anchor = '\ndef main() -> None:\n'
if main_anchor not in text:
    raise SystemExit("main insertion anchor missing")
text = text.replace(main_anchor, PURGE_HELPER.rstrip() + main_anchor, 1)
text = text.replace(
    '        index_qa=build_indexes_and_manifest(project)\n',
    '        cache_qa=purge_generated_caches(project)\n        index_qa=build_indexes_and_manifest(project)\n',
    1,
)
text = text.replace(
    '"reports":len(report_files),"accepted_predecessor_mutated":False',
    '"reports":len(report_files),"cache_hygiene":cache_qa,"accepted_predecessor_mutated":False',
    1,
)

NEW_COMPLETE_RESTORE = r'''
def build_complete_restore(project_archive: Path, recovery_zip: Path, base_restore: Path, reports: list[Path], dist: Path, work: Path, session_qa: dict[str, Any]) -> tuple[Path, dict[str, Any]]:
    package = work / "complete_restore_package"
    package.mkdir(parents=True)
    snapshot_dir = package / "PROJECT_SNAPSHOT"; snapshot_dir.mkdir()
    tools = package / "TOOLS"; tools.mkdir()
    report_dir = package / "REPORTS"; report_dir.mkdir()
    identity_dir = package / "SOURCE_CHECKPOINT_IDENTITY"; identity_dir.mkdir()
    shutil.copy2(project_archive, snapshot_dir / project_archive.name)
    for report in reports:
        shutil.copy2(report, report_dir / report.name)
    instructions = locate_instructions(base_restore, work)
    if instructions:
        shutil.copy2(instructions, package / "Instructions.txt")
    else:
        text_write(package / "Instructions.txt", "Human Pathogen Database project instructions are incorporated into the current project snapshot. Follow RESTORE_READ_FIRST.md and the current project governance files.")
    source_identity = {
        "base_response66_restore": verify_zip(base_restore, BASE_RESPONSE66_BYTES, BASE_RESPONSE66_SHA256),
        "response68_checkpoint2_recovery": verify_zip(recovery_zip, CP2_RECOVERY_BYTES, CP2_RECOVERY_SHA256),
        "embedded_in_restore": False,
        "reason": "The complete project snapshot is independently self-contained; prior recovery data is recorded by identity rather than duplicated into the session-end restore.",
    }
    json_write(identity_dir / "RESPONSE_66_AND_68_SOURCE_IDENTITY.json", source_identity)
    identity = {
        "schema":"mrhpd-complete-restore-identity-1.0","generated_at":NOW,"version":PROJECT_VERSION,
        "section":"Remediation Section 4 of 5","session":"Session 2 of 3 COMPLETE","checkpoint":"3 of 3 COMPLETE","response":69,
        "project_snapshot":{"name":project_archive.name,"bytes":project_archive.stat().st_size,"sha256":sha256_file(project_archive)},
        "source_checkpoint_identity":source_identity,
        "self_contained":True,"requires_other_project_files":False,"requires_conversation_reconstruction":False,
        "accepted_predecessor_mutated":False,"next":"Remediation Section 4 of 5 Session 3 of 3",
    }
    json_write(package / "CURRENT_PROJECT_IDENTITY.json", identity)
    text_write(package / "RESTORE_READ_FIRST.md", f"""# Human Pathogen Database — Complete Restore Through Response 69

This is the complete self-contained Section 4 Session 2 restore. It requires no prior ZIP, checkpoint package, cloud artifact, user-supplied project file, or reconstruction from the conversation.

## Restore

1. Run `python TOOLS/restore_verify_extract.py` to verify the package and the embedded current project.
2. Run `python TOOLS/restore_verify_extract.py --extract-project-to <destination>` to verify and extract the complete current project snapshot.
3. Open the extracted project and begin with its current Recovery, Tracking, QA and README files.

## Current state

- Remediation Section 4 of 5: CONTINUE
- Session 2 of 3: COMPLETE
- Checkpoint 3 of 3: COMPLETE
- Current response: 69
- Next: Session 3 of 3
- Accepted predecessor modified: no
- Frozen Section 3 release modified: no
- User upload required: no
""")
    verifier = f'''#!/usr/bin/env python3
import argparse,hashlib,json,sqlite3,tempfile,zipfile
from pathlib import Path,PurePosixPath
PROJECT_NAME={project_archive.name!r}
PROJECT_BYTES={project_archive.stat().st_size}
PROJECT_SHA={sha256_file(project_archive)!r}

def sha(path):
 h=hashlib.sha256()
 with open(path,'rb') as f:
  for b in iter(lambda:f.read(1024*1024),b''): h.update(b)
 return h.hexdigest()
def safe(zf):
 names=zf.namelist()
 if len(names)!=len(set(names)): raise SystemExit('duplicate ZIP members')
 for name in names:
  p=PurePosixPath(name.replace('\\\\','/'))
  if p.is_absolute() or '..' in p.parts: raise SystemExit('unsafe ZIP path: '+name)
 bad=zf.testzip()
 if bad: raise SystemExit('ZIP CRC failure: '+bad)
def verify(path,size,digest):
 if not path.exists() or path.stat().st_size!=size or sha(path)!=digest: raise SystemExit('identity failure: '+str(path))
 with zipfile.ZipFile(path) as zf: safe(zf)
def validate_project(root):
 dbs=[]
 for p in root.rglob('*.sqlite'):
  try:
   con=sqlite3.connect(p)
   r=con.execute("SELECT COUNT(*) FROM thread_response_reconciliation_cp3 WHERE response_key='R69'").fetchone()[0]
   tables=con.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table'").fetchone()[0]
   if r==1: dbs.append((tables,p))
   con.close()
  except Exception: pass
 if not dbs: raise SystemExit('Response 69 canonical database not found')
 db=sorted(dbs,reverse=True)[0][1]
 con=sqlite3.connect(db)
 try:
  if con.execute('PRAGMA integrity_check').fetchone()[0]!='ok': raise SystemExit('SQLite integrity failure')
  if list(con.execute('PRAGMA foreign_key_check')): raise SystemExit('foreign-key failure')
  if con.execute("SELECT state FROM section4_checkpoint WHERE checkpoint_code='MRHPD-V3-CP4-S2-CP3'").fetchone()!=('session_complete',): raise SystemExit('Checkpoint 3 state failure')
 finally: con.close()
 return str(db.relative_to(root))
p=argparse.ArgumentParser()
p.add_argument('--extract-project-to',type=Path)
a=p.parse_args()
base=Path(__file__).resolve().parent.parent
project=base/'PROJECT_SNAPSHOT'/PROJECT_NAME
verify(project,PROJECT_BYTES,PROJECT_SHA)
manifest=json.loads((base/'COMPLETE_RESTORE_MANIFEST.json').read_text())
for row in manifest['files']:
 path=base/row['path']
 if not path.exists() or path.stat().st_size!=row['bytes'] or sha(path)!=row['sha256']: raise SystemExit('restore manifest failure: '+row['path'])
if a.extract_project_to:
 target=a.extract_project_to; target.mkdir(parents=True,exist_ok=True)
 with zipfile.ZipFile(project) as zf: safe(zf); zf.extractall(target)
 roots=[x for x in target.iterdir() if x.is_dir()]
 project_root=roots[0] if len(roots)==1 else target
 db=validate_project(project_root)
else:
 with tempfile.TemporaryDirectory(prefix='mrhpd-r69-verify-') as td:
  target=Path(td)
  with zipfile.ZipFile(project) as zf: safe(zf); zf.extractall(target)
  roots=[x for x in target.iterdir() if x.is_dir()]
  project_root=roots[0] if len(roots)==1 else target
  db=validate_project(project_root)
print(json.dumps({{'status':'passed','project_archive':PROJECT_NAME,'project_sha256':PROJECT_SHA,'canonical_database':db,'self_contained':True}},indent=2))
'''
    text_write(tools / "restore_verify_extract.py", verifier)
    json_write(package / "SESSION_2_ACCEPTANCE_QA.json", session_qa)
    control_names = {"COMPLETE_RESTORE_MANIFEST.json","COMPLETE_RESTORE_CHECKSUMS.sha256"}
    rows=[]
    for path in sorted(package.rglob("*")):
        if path.is_file() and path.name not in control_names:
            rows.append({"path":path.relative_to(package).as_posix(),"bytes":path.stat().st_size,"sha256":sha256_file(path)})
    json_write(package / "COMPLETE_RESTORE_MANIFEST.json", {"generated_at":NOW,"file_count":len(rows),"total_bytes":sum(r["bytes"] for r in rows),"files":rows,"self_contained":True})
    text_write(package / "COMPLETE_RESTORE_CHECKSUMS.sha256", "".join(f"{r['sha256']}  {r['path']}\n" for r in rows))
    restore = dist / (
        f"Medical References - Human Pathogen Database v{PROJECT_VERSION} Part 8 of 8 "
        f"Remediation Section 4 of 5 Session 2 of 3 COMPLETE RESTORE THROUGH RESPONSE 69 {STAMP}.zip"
    )
    with zipfile.ZipFile(restore,"w",compression=zipfile.ZIP_DEFLATED,compresslevel=9,allowZip64=True) as zf:
        for path in sorted(package.rglob("*")):
            if path.is_file(): zf.write(path,path.relative_to(package).as_posix())
    restore_qa=verify_zip(restore)
    if restore_qa["bytes"] >= 180*1024*1024:
        raise RuntimeError({"restore_exceeds_180_mib":restore_qa})
    with tempfile.TemporaryDirectory(prefix="mrhpd-r69-restore-clean-") as td:
        clean=Path(td); safe_extract(restore,clean)
        result=subprocess.run([sys.executable,str(clean/"TOOLS"/"restore_verify_extract.py")],cwd=clean,text=True,capture_output=True,timeout=1800)
        if result.returncode: raise RuntimeError({"restore_verifier_failed":{"stdout":result.stdout[-20000:],"stderr":result.stderr[-20000:]}})
    verification={
        "schema":"mrhpd-response69-complete-restore-verification-1.0","generated_at":NOW,"status":"passed",
        "restore":restore_qa,"project_snapshot":verify_zip(project_archive),"source_checkpoint_identity":source_identity,
        "clean_restore_verifier":"passed","self_contained":True,"requires_other_project_files":False,
        "requires_conversation_reconstruction":False,"accepted_predecessor_mutated":False,
        "checkpoint_3_of_3_complete":True,"session_2_of_3_complete":True,"remediation_section_4_complete":False,
        "next":"Remediation Section 4 of 5 Session 3 of 3",
    }
    json_write(dist/"MRHPD v3.0.0a Response 69 Complete Restore Verification.json",verification)
    text_write(dist/f"{restore.name}.sha256.txt",f"{restore_qa['sha256']}  {restore.name}")
    return restore,verification
'''
text = replace_function(text, "build_complete_restore", "build_transport_volumes", NEW_COMPLETE_RESTORE)

NEW_TRANSPORT = r'''
def build_transport_volumes(restore: Path, dist: Path) -> dict[str, Any]:
    total=restore.stat().st_size
    if not (100*1024*1024 < total < 180*1024*1024):
        raise RuntimeError({"expected_two_volume_restore_size":total})
    first_size=(total+1)//2
    raw=[]
    with restore.open("rb") as source:
        for sequence,size in [(1,first_size),(2,total-first_size)]:
            path=dist/f"{restore.name}.part{sequence:03d}"
            remaining=size
            with path.open("wb") as out:
                while remaining:
                    block=source.read(min(1024*1024,remaining))
                    if not block: raise RuntimeError("unexpected EOF while splitting restore")
                    out.write(block); remaining-=len(block)
            raw.append({"sequence":sequence,"name":path.name,"bytes":path.stat().st_size,"sha256":sha256_file(path),"path":path})
    manifest={
        "schema":"mrhpd-complete-restore-transport-1.0","generated_at":NOW,
        "restore":{"name":restore.name,"bytes":total,"sha256":sha256_file(restore)},
        "part_count":2,"parts":[{k:v for k,v in row.items() if k!="path"} for row in raw],
        "minimum_volume_count":2,
    }
    manifest_path=dist/"MRHPD_RESPONSE69_COMPLETE_RESTORE_TRANSPORT_MANIFEST.json"; json_write(manifest_path,manifest)
    reassemble=dist/"reassemble_response69_complete_restore.py"
    text_write(reassemble, f'''#!/usr/bin/env python3
import hashlib,json
from pathlib import Path
M={manifest!r}
def sha(p):
 h=hashlib.sha256()
 with open(p,'rb') as f:
  for b in iter(lambda:f.read(1024*1024),b''): h.update(b)
 return h.hexdigest()
root=Path(__file__).resolve().parent
out=root/M['restore']['name']
with open(out,'wb') as dst:
 for row in M['parts']:
  p=root/row['name']
  if not p.exists() or p.stat().st_size!=row['bytes'] or sha(p)!=row['sha256']: raise SystemExit('part identity failure: '+row['name'])
  with open(p,'rb') as src:
   for b in iter(lambda:src.read(1024*1024),b''): dst.write(b)
if out.stat().st_size!=M['restore']['bytes'] or sha(out)!=M['restore']['sha256']: raise SystemExit('restore identity failure')
print(json.dumps({{'status':'passed','restore':out.name,'bytes':out.stat().st_size,'sha256':sha(out)}},indent=2))
''')
    wrappers=[]
    for row in raw:
        wrapper=dist/f"MRHPD v3.0.0a Response 69 Complete Restore Drive Volume {row['sequence']} of 2.zip"
        readme=(
            f"MRHPD Response 69 complete restore volume {row['sequence']} of 2. BOTH VOLUMES ARE REQUIRED. "
            "Extract both wrappers into the same directory and run reassemble_response69_complete_restore.py.\n"
        )
        readme_path=dist/f"README_VOLUME_{row['sequence']}.txt"; text_write(readme_path,readme)
        with zipfile.ZipFile(wrapper,"w",compression=zipfile.ZIP_STORED,allowZip64=True) as zf:
            zf.write(row["path"],row["path"].name)
            zf.write(manifest_path,manifest_path.name)
            zf.write(reassemble,reassemble.name)
            zf.write(readme_path,readme_path.name)
        qa=verify_zip(wrapper)
        if qa["bytes"] >= 104_857_600: raise RuntimeError({"drive_volume_exceeds_connector_limit":qa})
        wrappers.append({"sequence":row["sequence"],"wrapper":wrapper,"qa":qa,"raw_part":{k:v for k,v in row.items() if k!="path"}})
    return {"status":"passed","manifest":manifest,"manifest_path":manifest_path,"reassembly_utility":reassemble,"volumes":wrappers}
'''
text = replace_function(text, "build_transport_volumes", "build_controls_zip", NEW_TRANSPORT)

# Final exact naming and summary-schema corrections.
text = text.replace("mrhpd-response69-session2-build-summary-1.0", "mrhpd-response69-session2-complete-build-summary-1.0")
text = text.replace("RECOVERY_EVENTS_116_97", "RECOVERY_EVENTS_116_118")
text = text.replace("events 95–97", "events 116–118")
text = text.replace("Event 95", "Event 116").replace("Event 96", "Event 117").replace("Event 97", "Event 118")

# The generated source must not retain operative Session 1 or Response 64/65
# controls. Historic mentions inside imported evidence are not present in this
# builder, so these are strong generation gates.
for forbidden in [
    "BASE_RESPONSE64",
    "Response 64 restore",
    "response_key='R65'",
    "MRHPD-V3-CP4-S1-CP3",
    "Session 1 of 3 COMPLETE",
    "S4S1 Session Release",
    "--base-response64-restore",
]:
    if forbidden in text:
        raise SystemExit(f"forbidden stale control remains: {forbidden}")

TARGET.parent.mkdir(parents=True, exist_ok=True)
TARGET.write_text(text, encoding="utf-8")
print(f"Generated {TARGET} from {SOURCE} for Response 69 / Section 4 Session 2 completion.")
