#!/usr/bin/env python3
"""Build the MRHPD Section 4 Session 1 Checkpoint 2 recovery delta.

The builder starts from the independently verified complete restore through
Response 64, extracts the frozen final Section 3 project, creates a separate
mutable working tree, synchronizes the canonical SQLite database, comprehensive
workbook, local application and current tracking state, runs regressions, and
emits a deterministic checkpoint-recovery overlay. The frozen Section 3 release
and accepted predecessor are never edited in place.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import textwrap
import zipfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Any, Iterable

PROJECT_VERSION = "3.0.0a"
BASE_RESPONSE64_BYTES = 145_920_215
BASE_RESPONSE64_SHA256 = "71e7a06868e82238188827fd73ca7b2843670b0843dcf98ab5ab72305bf77834"
FINAL_SECTION3_BYTES = 147_057_203
FINAL_SECTION3_SHA256 = "2f517e809f49a30808a98491feb19aad20af557eb152db9fbae8603ef70fb402"
PUBLICATION_SHA256 = "8a053112ca24cd730b970130d5d0fc57a15c681531603601096186aeb0cd9642"
EDITABLE_ASSEMBLY_SHA256 = "f832ff934d77049d75712f28bdfc9167b8a6b119c797235431b304b9e24369a2"
RESPONSE_NUMBER = 65
NOW_DT = datetime.now(timezone.utc)
NOW = NOW_DT.replace(microsecond=0).isoformat().replace("+00:00", "Z")
STAMP = NOW_DT.strftime("%Y-%m-%d %H%M UTC")

RAW_PROMPTS = [
    {
        "prompt_number": "64.1",
        "prompt_text": (
            "Continue\n\nEmit the complete restore at the end of every session, and section, and of course, "
            "once the project is complete you should emit the entire project"
        ),
        "answered_by_response": 65,
        "source_id": "CURRENT-CONVERSATION-R65",
        "source_path": "Current conversation turn",
        "status": "source_verified",
    },
    {
        "prompt_number": "64.2",
        "prompt_text": "In between those emissions of the full product, emit the checkpoint recovery data",
        "answered_by_response": 65,
        "source_id": "CURRENT-CONVERSATION-R65",
        "source_path": "Current conversation turn",
        "status": "source_verified",
    },
    {
        "prompt_number": "64.3",
        "prompt_text": "If you’re currently working, carry on with what you’re doing. Otherwise, please resume.",
        "answered_by_response": 65,
        "source_id": "CURRENT-CONVERSATION-R65",
        "source_path": "Current conversation turn",
        "status": "source_verified",
    },
]

RAW_PROMPT_COMBINED = "\n\n--- NEXT USER MESSAGE ---\n\n".join(row["prompt_text"] for row in RAW_PROMPTS)

RESPONSE65 = {
    "response_key": "R65",
    "response_number": 65,
    "response_label": "65",
    "branch_id": "mainline",
    "canonical_current": 1,
    "response_date": NOW,
    "major_topic": "Human Pathogen Database remediation",
    "title": "Section 4 workbook and local-application synchronization checkpoint",
    "goal": (
        "Resume Section 4 Session 1, synchronize the current workbook and local application, and implement the "
        "full-restore-at-session/section/project and checkpoint-recovery-between policy."
    ),
    "raw_prompt": RAW_PROMPT_COMBINED,
    "raw_response": "[PRE-EMISSION RESPONSE; final user-visible response is represented by the source-supported summary]",
    "summary": (
        "Created a mutable Section 4 working tree from the frozen Section 3 release; synchronized the canonical SQLite database, "
        "comprehensive workbook, local application, Raw/Net tracking, response reconciliation, recovery records and checkpoint "
        "emission policy; reran database and application regressions; and emitted a verified checkpoint-recovery overlay."
    ),
    "state": "checkpoint_complete_continue_required",
    "coverage": "three exact fractional raw prompts plus source-supported response summary",
    "fidelity_classification": "source_verified_fractional_prompts_and_summary",
    "source_id": "CURRENT-CONVERSATION-R65",
    "source_path": "Current conversation turn and Checkpoint 2 recovery package",
    "notes": (
        "The next checkpoint ends Section 4 Session 1 and must emit a complete self-contained restore. Intermediate checkpoints "
        "emit recovery data tied to the last complete restore."
    ),
}

EMISSION_POLICY = {
    "schema": "mrhpd-emission-policy-2.0",
    "effective_at_response": 65,
    "recorded_at": NOW,
    "full_product_emissions": {
        "end_of_each_session": True,
        "end_of_each_section": True,
        "project_completion": True,
        "contents": "complete current project/product plus all restore, instructions, tracking, recovery, manifests, hashes and verification tools",
    },
    "intermediate_checkpoint_emissions": {
        "required": True,
        "contents": "complete checkpoint recovery delta, exact base identity, deterministic apply/verify utility, QA evidence and current tracking",
        "may_reference_last_complete_restore": True,
    },
    "user_download_host": "Google Drive",
    "sandbox_is_controlling": False,
    "accepted_predecessor_mutation_permitted": False,
}


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(chunk_size), b""):
            h.update(block)
    return h.hexdigest()


def json_write(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8")


def text_write(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")


def csv_write(path: Path, rows: list[dict[str, Any]], fields: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fields is None:
        fields = list(rows[0]) if rows else []
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def safe_infos(zf: zipfile.ZipFile, files_only: bool = False) -> list[zipfile.ZipInfo]:
    infos = [i for i in zf.infolist() if not files_only or not i.is_dir()]
    names = [i.filename for i in infos]
    duplicates = [name for name, count in Counter(names).items() if count > 1]
    unsafe = []
    for name in names:
        p = PurePosixPath(name.replace("\\", "/"))
        if name.startswith(("/", "\\")) or re.match(r"^[A-Za-z]:", name) or ".." in p.parts:
            unsafe.append(name)
    if duplicates or unsafe:
        raise RuntimeError({"duplicate_members": duplicates[:30], "unsafe_members": unsafe[:30]})
    return infos


def verify_zip(path: Path, expected_bytes: int | None = None, expected_sha256: str | None = None) -> dict[str, Any]:
    if expected_bytes is not None and path.stat().st_size != expected_bytes:
        raise RuntimeError({"file": str(path), "expected_bytes": expected_bytes, "actual_bytes": path.stat().st_size})
    digest = sha256_file(path)
    if expected_sha256 is not None and digest != expected_sha256:
        raise RuntimeError({"file": str(path), "expected_sha256": expected_sha256, "actual_sha256": digest})
    with zipfile.ZipFile(path) as zf:
        infos = safe_infos(zf)
        bad = zf.testzip()
        if bad:
            raise RuntimeError({"zip_crc_error": bad, "file": str(path)})
        filler = [i.filename for i in infos if re.search(r"(^|/)(filler|padding|pad)(/|$)", i.filename, re.I)]
        if filler:
            raise RuntimeError({"filler_members": filler[:30]})
    return {
        "file": path.name,
        "bytes": path.stat().st_size,
        "sha256": digest,
        "members": len(infos),
        "crc": "passed",
        "duplicate_members": 0,
        "unsafe_paths": 0,
        "filler_members": 0,
    }


def safe_extract(path: Path, target: Path) -> None:
    target.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path) as zf:
        safe_infos(zf)
        zf.extractall(target)


def locate_response64_restore(input_root: Path) -> Path:
    candidates = sorted(input_root.rglob("*.zip"), key=lambda p: p.stat().st_size, reverse=True)
    for path in candidates:
        if path.stat().st_size == BASE_RESPONSE64_BYTES and sha256_file(path) == BASE_RESPONSE64_SHA256:
            return path
    for wrapper in candidates:
        try:
            with zipfile.ZipFile(wrapper) as zf:
                safe_infos(zf)
                for info in zf.infolist():
                    if info.is_dir() or info.file_size != BASE_RESPONSE64_BYTES:
                        continue
                    with tempfile.NamedTemporaryFile(prefix="mrhpd-r64-", suffix=".zip", delete=False) as tmp:
                        tmp_path = Path(tmp.name)
                        with zf.open(info) as src:
                            shutil.copyfileobj(src, tmp, 1024 * 1024)
                    if sha256_file(tmp_path) == BASE_RESPONSE64_SHA256:
                        return tmp_path
                    tmp_path.unlink(missing_ok=True)
        except zipfile.BadZipFile:
            continue
    raise FileNotFoundError({
        "expected_bytes": BASE_RESPONSE64_BYTES,
        "expected_sha256": BASE_RESPONSE64_SHA256,
        "candidate_count": len(candidates),
        "candidate_sample": [(p.name, p.stat().st_size) for p in candidates[:20]],
    })


def find_exact(root: Path, pattern: str, expected_sha: str | None = None) -> Path:
    matches = sorted(root.rglob(pattern))
    if expected_sha:
        matches = [p for p in matches if sha256_file(p) == expected_sha]
    if len(matches) != 1:
        raise RuntimeError({"pattern": pattern, "matches": [str(p) for p in matches]})
    return matches[0]


def extract_final_project(base_restore: Path, work: Path) -> tuple[Path, Path, dict[str, Any]]:
    restore_root = work / "response64_restore"
    safe_extract(base_restore, restore_root)
    final_archive = find_exact(restore_root, "*FINAL SECTION 3 RELEASE*.zip", FINAL_SECTION3_SHA256)
    final_qa = verify_zip(final_archive, FINAL_SECTION3_BYTES, FINAL_SECTION3_SHA256)
    immutable_extract = work / "immutable_section3"
    safe_extract(final_archive, immutable_extract)
    dirs = [p for p in immutable_extract.iterdir() if p.is_dir()]
    files = [p for p in immutable_extract.iterdir() if p.is_file()]
    immutable_root = dirs[0] if len(dirs) == 1 and not files else immutable_extract
    mutable_parent = work / "mutable_section4"
    mutable_parent.mkdir(parents=True, exist_ok=True)
    mutable_root = mutable_parent / immutable_root.name
    shutil.copytree(immutable_root, mutable_root)
    return immutable_root, mutable_root, {
        "base_restore": verify_zip(base_restore, BASE_RESPONSE64_BYTES, BASE_RESPONSE64_SHA256),
        "final_section3_archive": final_qa,
        "immutable_project_root": immutable_root.name,
        "accepted_predecessor_mutated": False,
    }


def table_exists(con: sqlite3.Connection, table: str) -> bool:
    return con.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone() is not None


def table_columns(con: sqlite3.Connection, table: str) -> list[str]:
    return [row[1] for row in con.execute(f'PRAGMA table_info("{table}")')]


def upsert_by_unique(con: sqlite3.Connection, table: str, row: dict[str, Any], key: str) -> None:
    columns = table_columns(con, table)
    values = {name: value for name, value in row.items() if name in columns and not name.endswith("_id")}
    if key not in values:
        raise KeyError({"table": table, "required_key": key, "available": sorted(values)})
    names = list(values)
    update_names = [name for name in names if name != key]
    quoted = ", ".join(f'"{name}"' for name in names)
    placeholders = ", ".join("?" for _ in names)
    updates = ", ".join(f'"{name}"=excluded."{name}"' for name in update_names)
    con.execute(
        f'INSERT INTO "{table}" ({quoted}) VALUES ({placeholders}) '
        f'ON CONFLICT("{key}") DO UPDATE SET {updates}',
        [values[name] for name in names],
    )


def load_response64_progress(base_restore_root: Path) -> tuple[dict[str, Any] | None, list[dict[str, Any]]]:
    response_files = sorted(base_restore_root.rglob("Response_64_Tracking.json"))
    event_files = sorted(base_restore_root.rglob("RECOVERY_EVENTS_85_87.json"))
    response = None
    if response_files:
        response = json.loads(response_files[-1].read_text(encoding="utf-8-sig"))
    events: list[dict[str, Any]] = []
    if event_files:
        value = json.loads(event_files[-1].read_text(encoding="utf-8-sig"))
        if isinstance(value, list):
            events = value
    return response, events


def normalize_response64(record: dict[str, Any] | None) -> dict[str, Any]:
    if record:
        return {
            "response_key": "R64",
            "response_number": 64,
            "response_label": "64",
            "branch_id": "mainline",
            "canonical_current": 1,
            "response_date": record.get("created_at", "2026-07-31T03:55:00Z"),
            "major_topic": record.get("major_topic", "Human Pathogen Database remediation"),
            "title": record.get("title", "Persistent Google Drive download remediation and Section 4 intake"),
            "goal": record.get("goal"),
            "raw_prompt": record.get("raw_prompt"),
            "raw_response": record.get("raw_response"),
            "summary": record.get("summary"),
            "state": record.get("state", "checkpoint_complete_continue_required"),
            "coverage": record.get("coverage", "exact raw prompt plus source-supported summary"),
            "fidelity_classification": "source_verified_prompt_and_summary",
            "source_id": "R64-RESTORE",
            "source_path": "Complete Restore Through Response 64",
            "notes": "Response 64 was synchronized into the copied Section 4 database during Checkpoint 2.",
        }
    return {
        "response_key": "R64",
        "response_number": 64,
        "response_label": "64",
        "branch_id": "mainline",
        "canonical_current": 1,
        "response_date": "2026-07-31T03:55:00Z",
        "major_topic": "Human Pathogen Database remediation",
        "title": "Persistent Google Drive download remediation and Section 4 intake",
        "goal": "Correct the expired temporary download, verify durable Google Drive custody, and begin Section 4 intake.",
        "raw_prompt": "It says the file expired when I turned to download a zip. That should not be possible if you are starring it within Google Drive, right?\n\nContinue",
        "raw_response": None,
        "summary": "Identified the expired item as the temporary sandbox attachment, verified persistent Drive custody, rebuilt the complete restore through Response 64, and completed the Section 4 intake audit.",
        "state": "checkpoint_complete_continue_required",
        "coverage": "source-supported prompt and summary",
        "fidelity_classification": "source_verified_prompt_and_summary",
        "source_id": "R64-RECOVERED",
        "source_path": "Complete Restore Through Response 64",
        "notes": None,
    }


def checkpoint_events(base_events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    events = list(base_events)
    events.extend([
        {
            "event_number": 88,
            "event_code": "V3-CP4-S1-REC-RESTORE-AND-CHECKPOINT-EMISSION-POLICY-SYNCHRONIZED",
            "occurred_at": NOW,
            "failed_step": "None; emission policy synchronization completed.",
            "exact_error_or_reason": "The user clarified that complete restores are emitted at session, section and project boundaries, while intermediate turns emit checkpoint recovery data.",
            "intact_artifacts": "Complete Restore Through Response 64, frozen Section 3 release, accepted predecessor, Drive custody and all prior checkpoint data.",
            "recovery_action": "Stored the operative policy in SQLite, workbook, tracking files, recovery records and the Checkpoint 2 package.",
            "validation_result": "Policy present in all current governance layers.",
            "data_quality_effect": "None.",
            "next_checkpoint": "Section 4 Session 1 Checkpoint 3 emits the complete session restore.",
        },
        {
            "event_number": 89,
            "event_code": "V3-CP4-S1-REC-TEMPORARY-REMOTE-EXECUTION-LANE-USED",
            "occurred_at": NOW,
            "failed_step": "Start the local container or Python execution lane.",
            "exact_error_or_reason": "The local execution surfaces returned InvalidArgumentError before code startup.",
            "intact_artifacts": "All Google Drive source, restore and checkpoint files remained intact.",
            "recovery_action": "Used an isolated noncontrolling execution host, downloaded only the verified Response 64 artifact, and prepared outputs for immediate Google Drive persistence.",
            "validation_result": "No external host is required for restore; the emitted checkpoint recovery ZIP contains every delta and deterministic recovery tool.",
            "data_quality_effect": "None.",
            "next_checkpoint": "Continue from the Drive-hosted recovery package.",
        },
        {
            "event_number": 90,
            "event_code": "V3-CP4-S1-REC-WORKBOOK-APPLICATION-DATABASE-SYNCHRONIZED",
            "occurred_at": NOW,
            "failed_step": "None; Checkpoint 2 synchronization completed.",
            "exact_error_or_reason": "Section 4 required a copied-tree synchronization of the database, workbook, application and cumulative tracking state.",
            "intact_artifacts": "Frozen Section 3 release and accepted predecessor remain byte-identical and unmodified.",
            "recovery_action": "Created a new canonical Section 4 database, updated the workbook without deleting accepted sheets, pointed a copied application at the current database, reran regressions and built a deterministic overlay.",
            "validation_result": "Passed when SQLite, workbook, application, publication and overlay verification gates all succeed.",
            "data_quality_effect": "None; only a copied Section 4 working tree changed.",
            "next_checkpoint": "Checkpoint 3 of 3 completes Session 1 and emits a full self-contained restore.",
        },
    ])
    dedup: dict[str, dict[str, Any]] = {}
    for event in events:
        code = event.get("event_code")
        if code:
            dedup[code] = event
    return sorted(dedup.values(), key=lambda e: int(e.get("event_number") or e.get("remediation_recovery_event_id") or 10**9))


def apply_database_sync(mutable_root: Path, response64: dict[str, Any], base_events: list[dict[str, Any]]) -> tuple[Path, dict[str, Any], list[dict[str, Any]]]:
    dbs = sorted((mutable_root / "Database").glob("*.sqlite"), key=lambda p: p.stat().st_size, reverse=True)
    if not dbs:
        dbs = sorted(mutable_root.rglob("*.sqlite"), key=lambda p: p.stat().st_size, reverse=True)
    if not dbs:
        raise FileNotFoundError("No SQLite database found")
    source_db = dbs[0]
    target_name = (
        f"Medical References - Human Pathogen Database v{PROJECT_VERSION} Part 8 of 8 "
        "Remediation Section 4 of 5 Session 1 of 3 Checkpoint 2 of 3.sqlite"
    )
    target_db = source_db.with_name(target_name)
    shutil.copy2(source_db, target_db)
    events = checkpoint_events(base_events)

    con = sqlite3.connect(target_db)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys=ON")
    try:
        con.execute("BEGIN IMMEDIATE")
        if not table_exists(con, "thread_response_reconciliation_cp3"):
            raise RuntimeError("thread_response_reconciliation_cp3 table is missing")
        for response in [normalize_response64(response64), RESPONSE65]:
            row = dict(response)
            row["reconciled_at"] = NOW
            upsert_by_unique(con, "thread_response_reconciliation_cp3", row, "response_key")

        if not table_exists(con, "fractional_prompt_cp3"):
            raise RuntimeError("fractional_prompt_cp3 table is missing")
        for prompt in RAW_PROMPTS:
            row = dict(prompt)
            row["reconciled_at"] = NOW
            upsert_by_unique(con, "fractional_prompt_cp3", row, "prompt_number")

        if table_exists(con, "remediation_recovery_event"):
            columns = table_columns(con, "remediation_recovery_event")
            for event in events:
                normalized = {
                    "event_code": event.get("event_code"),
                    "occurred_at": event.get("occurred_at", NOW),
                    "failed_step": event.get("failed_step", "Not recorded"),
                    "exact_error_or_reason": event.get("exact_error_or_reason", "Not recorded"),
                    "intact_artifacts": event.get("intact_artifacts", "Not recorded"),
                    "recovery_action": event.get("recovery_action", "Not recorded"),
                    "validation_result": event.get("validation_result", "Not recorded"),
                    "data_quality_effect": event.get("data_quality_effect", "None"),
                    "next_checkpoint": event.get("next_checkpoint", "Checkpoint 3 of 3"),
                }
                normalized = {k: (json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
                              for k, v in normalized.items() if k in columns}
                upsert_by_unique(con, "remediation_recovery_event", normalized, "event_code")

        con.executescript("""
        CREATE TABLE IF NOT EXISTS restore_emission_policy (
          restore_emission_policy_id INTEGER PRIMARY KEY,
          policy_version TEXT NOT NULL UNIQUE,
          effective_response INTEGER NOT NULL,
          full_restore_at_session_end INTEGER NOT NULL CHECK(full_restore_at_session_end IN (0,1)),
          full_restore_at_section_end INTEGER NOT NULL CHECK(full_restore_at_section_end IN (0,1)),
          full_project_at_completion INTEGER NOT NULL CHECK(full_project_at_completion IN (0,1)),
          checkpoint_recovery_between INTEGER NOT NULL CHECK(checkpoint_recovery_between IN (0,1)),
          user_download_host TEXT NOT NULL,
          policy_json TEXT NOT NULL,
          recorded_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS section4_checkpoint (
          section4_checkpoint_id INTEGER PRIMARY KEY,
          checkpoint_code TEXT NOT NULL UNIQUE,
          section_label TEXT NOT NULL,
          session_label TEXT NOT NULL,
          checkpoint_label TEXT NOT NULL,
          response_number INTEGER NOT NULL,
          state TEXT NOT NULL,
          database_integrity TEXT,
          foreign_key_violations INTEGER,
          workbook_status TEXT,
          application_status TEXT,
          publication_sha256 TEXT,
          accepted_predecessor_mutated INTEGER NOT NULL CHECK(accepted_predecessor_mutated IN (0,1)),
          recorded_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS section4_sync_qa (
          section4_sync_qa_id INTEGER PRIMARY KEY,
          checkpoint_code TEXT NOT NULL,
          control_name TEXT NOT NULL,
          expected_value TEXT,
          actual_value TEXT,
          status TEXT NOT NULL,
          evidence_path TEXT,
          recorded_at TEXT NOT NULL,
          UNIQUE(checkpoint_code, control_name)
        );
        """)
        con.execute("""
          INSERT INTO restore_emission_policy
          (policy_version,effective_response,full_restore_at_session_end,full_restore_at_section_end,
           full_project_at_completion,checkpoint_recovery_between,user_download_host,policy_json,recorded_at)
          VALUES (?,?,?,?,?,?,?,?,?)
          ON CONFLICT(policy_version) DO UPDATE SET
            effective_response=excluded.effective_response,
            full_restore_at_session_end=excluded.full_restore_at_session_end,
            full_restore_at_section_end=excluded.full_restore_at_section_end,
            full_project_at_completion=excluded.full_project_at_completion,
            checkpoint_recovery_between=excluded.checkpoint_recovery_between,
            user_download_host=excluded.user_download_host,
            policy_json=excluded.policy_json,
            recorded_at=excluded.recorded_at
        """, (
            EMISSION_POLICY["schema"], 65, 1, 1, 1, 1, "Google Drive",
            json.dumps(EMISSION_POLICY, ensure_ascii=False), NOW,
        ))

        if table_exists(con, "metadata") and {"key", "value"}.issubset(table_columns(con, "metadata")):
            metadata_updates = {
                "version": PROJECT_VERSION,
                "current_remediation_section": "Remediation Section 4 of 5",
                "current_session": "Session 1 of 3",
                "current_checkpoint": "Checkpoint 2 of 3",
                "current_response": "65",
                "current_canonical_database": target_name,
                "restore_emission_policy": "full at session/section/project completion; checkpoint recovery between",
                "accepted_predecessor_mutated": "no",
                "last_updated_utc": NOW,
                "next_checkpoint": "Remediation Section 4 of 5 Session 1 of 3 Checkpoint 3 of 3",
            }
            for key, value in metadata_updates.items():
                con.execute("""
                  INSERT INTO metadata(key,value) VALUES (?,?)
                  ON CONFLICT(key) DO UPDATE SET value=excluded.value
                """, (key, value))

        integrity = con.execute("PRAGMA integrity_check").fetchone()[0]
        foreign_keys = [tuple(row) for row in con.execute("PRAGMA foreign_key_check")]
        if integrity != "ok" or foreign_keys:
            raise RuntimeError({"integrity": integrity, "foreign_keys": foreign_keys[:30]})

        response_rows = con.execute("SELECT COUNT(*) FROM thread_response_reconciliation_cp3").fetchone()[0]
        response65_rows = con.execute("SELECT COUNT(*) FROM thread_response_reconciliation_cp3 WHERE response_key='R65'").fetchone()[0]
        fractional_rows = con.execute("SELECT COUNT(*) FROM fractional_prompt_cp3").fetchone()[0]
        current_fractional = con.execute(
            "SELECT COUNT(*) FROM fractional_prompt_cp3 WHERE prompt_number IN ('64.1','64.2','64.3')"
        ).fetchone()[0]
        recovery_rows = con.execute("SELECT COUNT(*) FROM remediation_recovery_event").fetchone()[0] if table_exists(con, "remediation_recovery_event") else None
        table_count = con.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table'").fetchone()[0]

        con.execute("""
          INSERT INTO section4_checkpoint
          (checkpoint_code,section_label,session_label,checkpoint_label,response_number,state,
           database_integrity,foreign_key_violations,workbook_status,application_status,publication_sha256,
           accepted_predecessor_mutated,recorded_at)
          VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
          ON CONFLICT(checkpoint_code) DO UPDATE SET
            response_number=excluded.response_number,state=excluded.state,
            database_integrity=excluded.database_integrity,
            foreign_key_violations=excluded.foreign_key_violations,
            recorded_at=excluded.recorded_at
        """, (
            "MRHPD-V3-CP4-S1-CP2", "Remediation Section 4 of 5", "Session 1 of 3", "Checkpoint 2 of 3",
            65, "database_synchronized", integrity, len(foreign_keys), "pending", "pending", PUBLICATION_SHA256, 0, NOW,
        ))
        con.commit()
    except Exception:
        con.rollback()
        raise
    finally:
        con.close()

    qa = {
        "status": "passed",
        "source_database": source_db.relative_to(mutable_root).as_posix(),
        "canonical_database": target_db.relative_to(mutable_root).as_posix(),
        "bytes": target_db.stat().st_size,
        "sha256": sha256_file(target_db),
        "integrity": "ok",
        "foreign_key_violations": 0,
        "table_count": table_count,
        "response_records": response_rows,
        "response65_records": response65_rows,
        "fractional_prompt_records": fractional_rows,
        "current_fractional_prompt_records": current_fractional,
        "recovery_event_records": recovery_rows,
        "accepted_predecessor_mutated": False,
    }
    if response65_rows != 1 or current_fractional != 3:
        raise RuntimeError({"database_sync_counts": qa})
    return target_db, qa, events


def autosize_sheet(ws, max_width: int = 70) -> None:
    from openpyxl.styles import Alignment
    for column in ws.columns:
        letter = column[0].column_letter
        width = max(12, min(max_width, max(len(str(cell.value or "")) for cell in column[:250]) + 2))
        ws.column_dimensions[letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def update_workbook(mutable_root: Path, db: Path, database_qa: dict[str, Any], events: list[dict[str, Any]]) -> tuple[Path, dict[str, Any]]:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    candidates = sorted(mutable_root.rglob("*.xlsx"), key=lambda p: p.stat().st_size, reverse=True)
    if not candidates:
        raise FileNotFoundError("No workbook found")
    source = candidates[0]
    target = source.with_name(
        f"Medical References - Human Pathogen Database v{PROJECT_VERSION} Part 8 of 8 "
        "Remediation Section 4 of 5 Session 1 of 3 Checkpoint 2 of 3 Comprehensive Tracking.xlsx"
    )
    shutil.copy2(source, target)
    wb = load_workbook(target)
    original_sheets = list(wb.sheetnames)
    managed = [
        "S4S1 Dashboard", "S4S1 Responses", "S4S1 Fractional", "S4S1 Recovery",
        "S4S1 Database QA", "S4S1 Application QA", "S4S1 Emission Policy",
    ]
    for name in managed:
        if name in wb.sheetnames:
            del wb[name]

    navy, teal, gold, white, pale = "17324F", "167D86", "D4A928", "FFFFFF", "EAF4F4"
    dashboard = wb.create_sheet("S4S1 Dashboard", 0)
    dashboard.merge_cells("A1:D2")
    dashboard["A1"] = "Human Pathogen Database — Section 4 Session 1 Checkpoint 2"
    dashboard["A1"].font = Font(bold=True, color=white, size=16)
    dashboard["A1"].fill = PatternFill("solid", fgColor=navy)
    dashboard["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    dashboard.append(["Control", "Expected", "Actual", "Status"])
    dashboard_rows = [
        ["Frozen final Section 3 release", FINAL_SECTION3_SHA256, FINAL_SECTION3_SHA256, "PASS"],
        ["Accepted predecessor mutation", "No", "No", "PASS"],
        ["Canonical database integrity", "ok", database_qa["integrity"], "PASS"],
        ["Foreign-key violations", 0, database_qa["foreign_key_violations"], "PASS"],
        ["Response 65 records", 1, database_qa["response65_records"], "PASS"],
        ["Current fractional prompts", 3, database_qa["current_fractional_prompt_records"], "PASS"],
        ["Application regression", "All tests pass", "Pending application phase", "PENDING"],
        ["Session-end complete restore", "Checkpoint 3", "Not due at intermediate Checkpoint 2", "PENDING"],
    ]
    for row in dashboard_rows:
        dashboard.append(row)
    for cell in dashboard[3]:
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=teal)
    dashboard.freeze_panes = "A4"
    dashboard.auto_filter.ref = f"A3:D{dashboard.max_row}"

    responses = wb.create_sheet("S4S1 Responses")
    response_headers = [
        "response_key", "response_number", "title", "goal", "raw_prompt", "summary",
        "state", "coverage", "fidelity_classification", "source_path",
    ]
    responses.append(response_headers)
    responses.append([RESPONSE65.get(h) for h in response_headers])

    fractional = wb.create_sheet("S4S1 Fractional")
    fractional_headers = ["prompt_number", "prompt_text", "answered_by_response", "source_id", "source_path", "status"]
    fractional.append(fractional_headers)
    for row in RAW_PROMPTS:
        fractional.append([row.get(h) for h in fractional_headers])

    recovery = wb.create_sheet("S4S1 Recovery")
    event_headers = [
        "event_number", "event_code", "occurred_at", "failed_step", "exact_error_or_reason",
        "intact_artifacts", "recovery_action", "validation_result", "data_quality_effect", "next_checkpoint",
    ]
    recovery.append(event_headers)
    for event in events:
        if int(event.get("event_number") or 0) >= 85:
            recovery.append([event.get(h) for h in event_headers])

    dbws = wb.create_sheet("S4S1 Database QA")
    dbws.append(["Control", "Value"])
    for key, value in database_qa.items():
        dbws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])

    appws = wb.create_sheet("S4S1 Application QA")
    appws.append(["Control", "Value"])
    appws.append(["Canonical database", db.relative_to(mutable_root).as_posix()])
    appws.append(["Application test state", "Pending application execution phase"])

    policy = wb.create_sheet("S4S1 Emission Policy")
    policy.append(["Policy area", "Operative requirement"])
    policy_rows = [
        ["End of each session", "Emit complete self-contained restore"],
        ["End of each section", "Emit complete self-contained restore"],
        ["Project completion", "Emit entire final project"],
        ["Between full emissions", "Emit checkpoint recovery data"],
        ["User-download host", "Google Drive"],
        ["Accepted predecessor", "Never modify in place"],
    ]
    for row in policy_rows:
        policy.append(row)

    for ws in [responses, fractional, recovery, dbws, appws, policy]:
        for cell in ws[1]:
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=navy)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        autosize_sheet(ws)
    autosize_sheet(dashboard)
    wb.save(target)

    # Independent reopen and formula-error scan.
    wb2 = load_workbook(target, data_only=False, read_only=True)
    error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}
    formula_errors: list[str] = []
    formula_count = 0
    for ws in wb2.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                if isinstance(value, str) and any(token in value for token in error_tokens):
                    formula_errors.append(f"{ws.title}!{cell.coordinate}:{value}")
                    if len(formula_errors) >= 100:
                        break
            if len(formula_errors) >= 100:
                break
        if len(formula_errors) >= 100:
            break
    final_sheets = list(wb2.sheetnames)
    wb2.close()
    missing_original = sorted(set(original_sheets) - set(final_sheets))
    missing_managed = sorted(set(managed) - set(final_sheets))
    qa = {
        "status": "passed" if not formula_errors and not missing_original and not missing_managed else "failed",
        "source_workbook": source.relative_to(mutable_root).as_posix(),
        "current_workbook": target.relative_to(mutable_root).as_posix(),
        "bytes": target.stat().st_size,
        "sha256": sha256_file(target),
        "source_sheet_count": len(original_sheets),
        "current_sheet_count": len(final_sheets),
        "original_sheets_preserved": not missing_original,
        "missing_original_sheets": missing_original,
        "managed_sheets": managed,
        "missing_managed_sheets": missing_managed,
        "formula_count": formula_count,
        "formula_error_count": len(formula_errors),
        "formula_errors": formula_errors,
    }
    if qa["status"] != "passed":
        raise RuntimeError({"workbook_qa": qa})
    return target, qa


def update_application(mutable_root: Path, db: Path) -> tuple[list[Path], dict[str, Any]]:
    app_files = sorted(mutable_root.rglob("human_pathogen_app.py"))
    if len(app_files) != 1:
        raise RuntimeError({"human_pathogen_app_matches": [str(p) for p in app_files]})
    app = app_files[0]
    original = app.read_text(encoding="utf-8")
    text = original
    sqlite_literals = re.findall(r"[^\"'\n]*Human Pathogen Database[^\"'\n]*\.sqlite", text)
    replacement_count = 0
    for literal in sorted(set(sqlite_literals), key=len, reverse=True):
        if literal in text:
            text = text.replace(literal, db.name)
            replacement_count += 1
    if db.name not in text:
        text = re.sub(r"([\"'])[^\"'\n]*\.sqlite\1", lambda m: f'{m.group(1)}{db.name}{m.group(1)}', text, count=1)
    text = text.replace("Remediation Section 3 of 5", "Remediation Section 4 of 5")
    text = text.replace("Session 4 of 4", "Session 1 of 3")
    app.write_text(text, encoding="utf-8")

    state = {
        "schema": "mrhpd-current-application-state-1.0",
        "generated_at": NOW,
        "remediation_section": "4 of 5",
        "session": "1 of 3",
        "checkpoint": "2 of 3",
        "response": 65,
        "canonical_database": db.name,
        "database_relative_path": db.relative_to(mutable_root).as_posix(),
        "accepted_predecessor_mutated": False,
        "restore_policy": EMISSION_POLICY,
    }
    state_path = app.parent / "CURRENT_PROJECT_STATE.json"
    json_write(state_path, state)
    readme_path = app.parent / "README_SECTION4_SESSION1_CHECKPOINT2.md"
    text_write(readme_path, f"""# Human Pathogen Database local application — Section 4 Session 1 Checkpoint 2

+Canonical database: `{db.name}`
+
+The application operates on the copied Section 4 database. The frozen Section 3 release and accepted predecessor remain immutable. Direct and loopback HTTP/security regressions are rerun during this checkpoint.
+""")
+
+    test_results: list[dict[str, Any]] = []
+    test_files = sorted(app.parent.glob("test*.py"))
+    if not test_files:
+        raise RuntimeError("No application regression tests found")
+    for test in test_files:
+        env = os.environ.copy()
+        env["MRHPD_DATABASE"] = str(db)
+        env["MRHPD_DB_PATH"] = str(db)
+        result = subprocess.run(
+            [sys.executable, str(test)], cwd=app.parent, env=env,
+            text=True, capture_output=True, timeout=300,
+        )
+        record = {
+            "test": test.name,
+            "returncode": result.returncode,
+            "stdout": result.stdout[-20000:],
+            "stderr": result.stderr[-20000:],
+        }
+        test_results.append(record)
+        if result.returncode != 0:
+            raise RuntimeError({"application_test_failed": record})
+
+    qa = {
+        "status": "passed",
+        "application": app.relative_to(mutable_root).as_posix(),
+        "application_sha256": sha256_file(app),
+        "database_name_replacement_count": replacement_count,
+        "canonical_database_referenced": db.name in app.read_text(encoding="utf-8"),
+        "state_file": state_path.relative_to(mutable_root).as_posix(),
+        "tests": test_results,
+        "test_count": len(test_results),
+        "all_returncodes_zero": all(item["returncode"] == 0 for item in test_results),
+    }
+    if not qa["canonical_database_referenced"]:
+        raise RuntimeError({"application_database_reference_missing": db.name})
+    return [app, state_path, readme_path, *test_files], qa
+
+
+def verify_publication(mutable_root: Path) -> dict[str, Any]:
+    from pypdf import PdfReader
+    pdfs = sorted(mutable_root.rglob("*Integrated Manuscript*.pdf"), key=lambda p: p.stat().st_size, reverse=True)
+    docxs = sorted(mutable_root.rglob("*Editable Integrated Manuscript Assembly*.docx"), key=lambda p: p.stat().st_size, reverse=True)
+    if not pdfs or not docxs:
+        raise RuntimeError({"pdfs": [str(p) for p in pdfs], "docxs": [str(p) for p in docxs]})
+    pdf, docx = pdfs[0], docxs[0]
+    pdf_sha, docx_sha = sha256_file(pdf), sha256_file(docx)
+    reader = PdfReader(str(pdf))
+    searchable = sum(1 for page in reader.pages if (page.extract_text() or "").strip())
+    qa = {
+        "status": "passed" if pdf_sha == PUBLICATION_SHA256 and docx_sha == EDITABLE_ASSEMBLY_SHA256 and len(reader.pages) == 537 and searchable == 537 else "failed",
+        "integrated_publication": pdf.relative_to(mutable_root).as_posix(),
+        "publication_bytes": pdf.stat().st_size,
+        "publication_sha256": pdf_sha,
+        "publication_pages": len(reader.pages),
+        "searchable_pages": searchable,
+        "editable_assembly": docx.relative_to(mutable_root).as_posix(),
+        "editable_assembly_bytes": docx.stat().st_size,
+        "editable_assembly_sha256": docx_sha,
+        "publication_unchanged": pdf_sha == PUBLICATION_SHA256,
+        "editable_assembly_unchanged": docx_sha == EDITABLE_ASSEMBLY_SHA256,
+    }
+    if qa["status"] != "passed":
+        raise RuntimeError({"publication_qa": qa})
+    return qa
+
+
+def build_tracking_and_governance(mutable_root: Path, database_qa: dict[str, Any], workbook_qa: dict[str, Any], app_qa: dict[str, Any], publication_qa: dict[str, Any], events: list[dict[str, Any]]) -> list[Path]:
+    target = mutable_root / "Tracking" / "Section 4 Session 1" / "Checkpoint 2"
+    target.mkdir(parents=True, exist_ok=True)
+    files: list[Path] = []
+    json_write(target / "Response_65_Tracking.json", RESPONSE65); files.append(target / "Response_65_Tracking.json")
+    json_write(target / "FRACTIONAL_PROMPTS_64_1_TO_64_3.json", RAW_PROMPTS); files.append(target / "FRACTIONAL_PROMPTS_64_1_TO_64_3.json")
+    json_write(target / "RECOVERY_EVENTS_85_90.json", events); files.append(target / "RECOVERY_EVENTS_85_90.json")
+    json_write(target / "RESTORE_AND_CHECKPOINT_EMISSION_POLICY.json", EMISSION_POLICY); files.append(target / "RESTORE_AND_CHECKPOINT_EMISSION_POLICY.json")
+
+    raw_net = f"""# Human Pathogen Database — Raw and Net Tracking Through Response 65
+
+## Major topic
+Human Pathogen Database remediation
+
+## Raw Prompt 64.1
+
+{RAW_PROMPTS[0]['prompt_text']}
+
+## Raw Prompt 64.2
+
+{RAW_PROMPTS[1]['prompt_text']}
+
+## Raw Prompt 64.3
+
+{RAW_PROMPTS[2]['prompt_text']}
+
+## Net Prompt through Response 65
+
+Continue the Human Pathogen Database from the complete restore through Response 64. Preserve Google Drive as the controlling user-download and project-recovery store. At the end of every session and section, emit a complete self-contained restore; at project completion, emit the entire project. Between those full-product emissions, emit complete checkpoint recovery data with exact base identity, all current deltas, deterministic application and verification tools, manifests, hashes, QA evidence, and current Raw/Net tracking. Complete Section 4 Session 1 Checkpoint 2 by synchronizing the copied SQLite database, comprehensive workbook, local application, and cumulative governance state without modifying the frozen Section 3 release or accepted predecessor.
+
+## Net Response through Response 65
+
+The frozen Section 3 release remains immutable. A copied Section 4 tree now contains the current canonical database, workbook, application state, response and fractional-prompt records, recovery events, emission policy, QA evidence, updated indexes and manifests. Database, workbook, application and publication gates passed. The emitted Checkpoint 2 recovery bundle reconstructs this state from the last full restore. Checkpoint 3 will complete Session 1 and emit the next complete self-contained restore.
+"""
+    text_write(target / "RAW_AND_NET_TRACKING.md", raw_net); files.append(target / "RAW_AND_NET_TRACKING.md")
+
+    index = f"""# Cumulative Thread Index Update — Response 65
+
+## Major topic
+Human Pathogen Database remediation
+
+## Response 65 — Section 4 workbook and local-application synchronization checkpoint
+
+**Goal:** Implement the clarified restore/checkpoint emission policy and complete Section 4 Session 1 Checkpoint 2 synchronization.
+
+**Output:** Created a separate mutable Section 4 project tree; synchronized Responses 64–65, fractional prompts 64.1–64.3, Recovery Events 85–90, the canonical SQLite database, comprehensive workbook and local application; preserved the 537-page publication and editable assembly byte-for-byte; reran regressions; and built a deterministic checkpoint-recovery overlay.
+
+**Disposition:** Remediation Section 4 of 5, Session 1 of 3, Checkpoint 2 of 3 COMPLETE. CONTINUE to Checkpoint 3 of 3, which ends Session 1 and emits a complete self-contained restore.
+"""
+    text_write(target / "CUMULATIVE_THREAD_INDEX_UPDATE.md", index); files.append(target / "CUMULATIVE_THREAD_INDEX_UPDATE.md")
+
+    addendum = """# Operative Restore and Checkpoint Emission Addendum — Response 65
+
+1. At the end of every session, emit a complete self-contained restore of the project current through that response.
+2. At the end of every section, emit a complete self-contained restore of the project current through that response.
+3. At final project completion, emit the entire project.
+4. Between those full-product emissions, emit checkpoint recovery data containing every current delta, exact base identity, deterministic apply/verify tools, manifests, checksums, QA evidence, recovery records and current Raw/Net tracking.
+5. Google Drive is the controlling user-download host.
+6. The frozen Section 3 release and accepted predecessor remain immutable.
+"""
+    text_write(target / "PROJECT_INSTRUCTIONS_ADDENDUM_RESPONSE65.md", addendum); files.append(target / "PROJECT_INSTRUCTIONS_ADDENDUM_RESPONSE65.md")
+
+    qa_dir = mutable_root / "QA" / "Section 4 Session 1" / "Checkpoint 2"
+    qa_dir.mkdir(parents=True, exist_ok=True)
+    qa_payloads = {
+        "DATABASE_QA.json": database_qa,
+        "WORKBOOK_QA.json": workbook_qa,
+        "APPLICATION_QA.json": app_qa,
+        "PUBLICATION_QA.json": publication_qa,
+    }
+    for name, payload in qa_payloads.items():
+        json_write(qa_dir / name, payload); files.append(qa_dir / name)
+
+    checkpoint = {
+        "schema": "mrhpd-recovery-checkpoint-1.0",
+        "project": "Medical References - Human Pathogen Database",
+        "version": PROJECT_VERSION,
+        "section": "Remediation Section 4 of 5",
+        "session": "Session 1 of 3",
+        "checkpoint": "2 of 3",
+        "response": 65,
+        "status": "COMPLETE",
+        "created_at": NOW,
+        "base_complete_restore": {"bytes": BASE_RESPONSE64_BYTES, "sha256": BASE_RESPONSE64_SHA256},
+        "frozen_section3_release": {"bytes": FINAL_SECTION3_BYTES, "sha256": FINAL_SECTION3_SHA256, "mutated": False},
+        "database": database_qa,
+        "workbook": workbook_qa,
+        "application": app_qa,
+        "publication": publication_qa,
+        "emission_policy": EMISSION_POLICY,
+        "next": "Remediation Section 4 of 5 Session 1 of 3 Checkpoint 3 of 3; complete session restore emission",
+    }
+    recovery_dir = mutable_root / "Recovery" / "Section 4 Session 1 Checkpoint 2"
+    recovery_dir.mkdir(parents=True, exist_ok=True)
+    json_write(recovery_dir / "CHECKPOINT_2_RECOVERY.json", checkpoint); files.append(recovery_dir / "CHECKPOINT_2_RECOVERY.json")
+    text_write(recovery_dir / "README.md", """# Section 4 Session 1 Checkpoint 2
+
+This directory records the copied-tree database, workbook and local-application synchronization. Checkpoint 3 of 3 will complete Session 1 and emit the next full self-contained restore.
+"""); files.append(recovery_dir / "README.md")
+    return files
+
+
+def build_project_indexes_and_manifest(mutable_root: Path) -> dict[str, Any]:
+    index_dir = mutable_root / "Indexes" / "Section 4 Session 1 Checkpoint 2"
+    manifest_dir = mutable_root / "Manifest" / "Section 4 Session 1 Checkpoint 2"
+    index_dir.mkdir(parents=True, exist_ok=True)
+    manifest_dir.mkdir(parents=True, exist_ok=True)
+    control_paths = {
+        (index_dir / "Source and Artifact Index.csv").relative_to(mutable_root).as_posix(),
+        (index_dir / "Source and Artifact Index.json").relative_to(mutable_root).as_posix(),
+        (index_dir / "Checkpoint Bit Index.sqlite").relative_to(mutable_root).as_posix(),
+        (manifest_dir / "Checkpoint Project Manifest.csv").relative_to(mutable_root).as_posix(),
+        (manifest_dir / "Checkpoint Project Manifest.json").relative_to(mutable_root).as_posix(),
+        (manifest_dir / "Checkpoint SHA256 Inventory.txt").relative_to(mutable_root).as_posix(),
+    }
+    physical = []
+    for path in sorted(mutable_root.rglob("*")):
+        if not path.is_file():
+            continue
+        rel = path.relative_to(mutable_root).as_posix()
+        if rel in control_paths:
+            continue
+        physical.append(path)
+    rows = []
+    text_extensions = {".txt", ".md", ".csv", ".json", ".py", ".html", ".css", ".js", ".xml", ".sql", ".yaml", ".yml"}
+    for path in physical:
+        rel = path.relative_to(mutable_root).as_posix()
+        rows.append({
+            "path": rel,
+            "name": path.name,
+            "extension": path.suffix.lower(),
+            "bytes": path.stat().st_size,
+            "sha256": sha256_file(path),
+            "top_level_category": rel.split("/", 1)[0],
+            "searchable": "yes" if path.suffix.lower() in text_extensions else "metadata_only",
+            "indexed_at": NOW,
+        })
+    csv_write(index_dir / "Source and Artifact Index.csv", rows)
+    json_write(index_dir / "Source and Artifact Index.json", {"generated_at": NOW, "file_count": len(rows), "files": rows})
+
+    bit_db = index_dir / "Checkpoint Bit Index.sqlite"
+    bit_db.unlink(missing_ok=True)
+    con = sqlite3.connect(bit_db)
+    try:
+        con.executescript("""
+        CREATE TABLE file_record(
+          file_record_id INTEGER PRIMARY KEY,
+          path TEXT NOT NULL UNIQUE,
+          name TEXT NOT NULL,
+          extension TEXT,
+          bytes INTEGER NOT NULL,
+          sha256 TEXT NOT NULL,
+          category TEXT NOT NULL,
+          indexed_at TEXT NOT NULL
+        );
+        CREATE VIRTUAL TABLE text_fts USING fts5(path UNINDEXED, content);
+        """)
+        for row, path in zip(rows, physical):
+            con.execute(
+                "INSERT INTO file_record(path,name,extension,bytes,sha256,category,indexed_at) VALUES (?,?,?,?,?,?,?)",
+                (row["path"], row["name"], row["extension"], row["bytes"], row["sha256"], row["top_level_category"], NOW),
+            )
+            if path.suffix.lower() in text_extensions:
+                content = path.read_text(encoding="utf-8", errors="replace")[:1_000_000]
+                if content:
+                    con.execute("INSERT INTO text_fts(path,content) VALUES (?,?)", (row["path"], content))
+        con.commit()
+        integrity = con.execute("PRAGMA integrity_check").fetchone()[0]
+        search_tests = {
+            term: con.execute("SELECT COUNT(*) FROM text_fts WHERE text_fts MATCH ?", (f'"{term}"',)).fetchone()[0]
+            for term in ["Response 65", "Checkpoint 2", "Google Drive", "emission policy"]
+        }
+    finally:
+        con.close()
+    if integrity != "ok" or any(value < 1 for value in search_tests.values()):
+        raise RuntimeError({"bit_index_integrity": integrity, "search_tests": search_tests})
+
+    csv_write(manifest_dir / "Checkpoint Project Manifest.csv", rows)
+    json_write(manifest_dir / "Checkpoint Project Manifest.json", {
+        "generated_at": NOW,
+        "file_count": len(rows),
+        "total_bytes": sum(row["bytes"] for row in rows),
+        "files": rows,
+        "recursive_controls_excluded": sorted(control_paths),
+    })
+    text_write(manifest_dir / "Checkpoint SHA256 Inventory.txt", "".join(f"{row['sha256']}  {row['path']}\n" for row in rows))
+    return {
+        "status": "passed",
+        "physical_file_count": len(rows),
+        "physical_bytes": sum(row["bytes"] for row in rows),
+        "source_index": (index_dir / "Source and Artifact Index.csv").relative_to(mutable_root).as_posix(),
+        "bit_index": bit_db.relative_to(mutable_root).as_posix(),
+        "bit_index_integrity": integrity,
+        "search_tests": search_tests,
+        "project_manifest": (manifest_dir / "Checkpoint Project Manifest.json").relative_to(mutable_root).as_posix(),
+    }
+
+
+def update_checkpoint_database_status(db: Path, workbook_qa: dict[str, Any], app_qa: dict[str, Any], index_qa: dict[str, Any]) -> None:
+    con = sqlite3.connect(db)
+    try:
+        con.execute("PRAGMA foreign_keys=ON")
+        con.execute("BEGIN IMMEDIATE")
+        con.execute("""
+          UPDATE section4_checkpoint SET
+            state='checkpoint_complete', workbook_status=?, application_status=?, recorded_at=?
+          WHERE checkpoint_code='MRHPD-V3-CP4-S1-CP2'
+        """, (workbook_qa["status"], app_qa["status"], NOW))
+        controls = [
+            ("SQLite integrity", "ok", "ok", "PASS", "QA/Section 4 Session 1/Checkpoint 2/DATABASE_QA.json"),
+            ("Foreign-key violations", "0", "0", "PASS", "QA/Section 4 Session 1/Checkpoint 2/DATABASE_QA.json"),
+            ("Workbook synchronization", "passed", workbook_qa["status"], "PASS", "QA/Section 4 Session 1/Checkpoint 2/WORKBOOK_QA.json"),
+            ("Application regressions", "passed", app_qa["status"], "PASS", "QA/Section 4 Session 1/Checkpoint 2/APPLICATION_QA.json"),
+            ("Publication unchanged", PUBLICATION_SHA256, PUBLICATION_SHA256, "PASS", "QA/Section 4 Session 1/Checkpoint 2/PUBLICATION_QA.json"),
+            ("Checkpoint Bit Index", "ok", index_qa["bit_index_integrity"], "PASS", index_qa["bit_index"]),
+        ]
+        for name, expected, actual, status, evidence in controls:
+            con.execute("""
+              INSERT INTO section4_sync_qa
+              (checkpoint_code,control_name,expected_value,actual_value,status,evidence_path,recorded_at)
+              VALUES (?,?,?,?,?,?,?)
+              ON CONFLICT(checkpoint_code,control_name) DO UPDATE SET
+                expected_value=excluded.expected_value,actual_value=excluded.actual_value,
+                status=excluded.status,evidence_path=excluded.evidence_path,recorded_at=excluded.recorded_at
+            """, ("MRHPD-V3-CP4-S1-CP2", name, expected, actual, status, evidence, NOW))
+        integrity = con.execute("PRAGMA integrity_check").fetchone()[0]
+        fk = list(con.execute("PRAGMA foreign_key_check"))
+        if integrity != "ok" or fk:
+            raise RuntimeError({"integrity": integrity, "foreign_keys": fk[:20]})
+        con.commit()
+    except Exception:
+        con.rollback()
+        raise
+    finally:
+        con.close()
+
+
+def build_report_docx(path: Path, qa: dict[str, Any]) -> None:
+    from docx import Document
+    from docx.enum.text import WD_ALIGN_PARAGRAPH
+    from docx.shared import Inches, Pt
+
+    doc = Document()
+    section = doc.sections[0]
+    section.top_margin = Inches(0.6)
+    section.bottom_margin = Inches(0.6)
+    section.left_margin = Inches(0.7)
+    section.right_margin = Inches(0.7)
+    doc.styles["Normal"].font.name = "Aptos"
+    doc.styles["Normal"].font.size = Pt(9)
+    title = doc.add_heading("Human Pathogen Database", 0)
+    title.alignment = WD_ALIGN_PARAGRAPH.CENTER
+    subtitle = doc.add_paragraph("Remediation Section 4 of 5 · Session 1 of 3 · Checkpoint 2 of 3")
+    subtitle.alignment = WD_ALIGN_PARAGRAPH.CENTER
+    doc.add_heading("Workbook and local-application synchronization", 1)
+    doc.add_paragraph(
+        "A separate mutable Section 4 tree was created from the frozen Section 3 release. The canonical SQLite database, comprehensive workbook, local application and current tracking were synchronized without modifying the accepted predecessor or frozen release."
+    )
+    table = doc.add_table(rows=1, cols=3)
+    table.style = "Table Grid"
+    for i, label in enumerate(["Control", "Result", "Evidence"]):
+        table.rows[0].cells[i].text = label
+    rows = [
+        ("Database integrity", qa["database"]["integrity"], qa["database"]["canonical_database"]),
+        ("Foreign-key violations", qa["database"]["foreign_key_violations"], "DATABASE_QA.json"),
+        ("Workbook status", qa["workbook"]["status"], qa["workbook"]["current_workbook"]),
+        ("Workbook sheets", qa["workbook"]["current_sheet_count"], "All accepted sheets preserved"),
+        ("Application status", qa["application"]["status"], qa["application"]["application"]),
+        ("Application tests", qa["application"]["test_count"], "All return codes zero"),
+        ("Publication pages", qa["publication"]["publication_pages"], "537 searchable pages"),
+        ("Publication unchanged", qa["publication"]["publication_unchanged"], qa["publication"]["publication_sha256"]),
+        ("Checkpoint Bit Index", qa["indexes"]["bit_index_integrity"], qa["indexes"]["bit_index"]),
+    ]
+    for control, result, evidence in rows:
+        cells = table.add_row().cells
+        cells[0].text, cells[1].text, cells[2].text = str(control), str(result), str(evidence)
+    doc.add_heading("Emission policy", 1)
+    doc.add_paragraph(
+        "Complete self-contained restores are emitted at the end of every session and section and at project completion. Intermediate turns emit complete checkpoint recovery data tied to the latest complete restore. Google Drive is the controlling download host."
+    )
+    doc.add_heading("Next checkpoint", 1)
+    doc.add_paragraph(
+        "Checkpoint 3 of 3 completes Section 4 Session 1, performs final session-level synchronization and QA, and emits the next complete self-contained restore."
+    )
+    doc.core_properties.title = "MRHPD Section 4 Session 1 Checkpoint 2 Recovery Report"
+    doc.core_properties.author = "Brent McAnulty, M.D."
+    doc.save(path)
+
+
+def build_report_pdf(path: Path, qa: dict[str, Any]) -> None:
+    from reportlab.lib import colors
+    from reportlab.lib.pagesizes import letter
+    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
+    from reportlab.lib.units import inch
+    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
+
+    styles = getSampleStyleSheet()
+    styles.add(ParagraphStyle(name="SmallHP", parent=styles["BodyText"], fontSize=8.5, leading=10.5))
+    story = [
+        Paragraph("Human Pathogen Database", styles["Title"]),
+        Paragraph("Remediation Section 4 of 5 · Session 1 of 3 · Checkpoint 2 of 3", styles["Heading2"]),
+        Paragraph("Workbook and local-application synchronization", styles["Heading1"]),
+        Paragraph(
+            "A separate mutable Section 4 tree was synchronized without modifying the accepted predecessor or frozen Section 3 release.",
+            styles["SmallHP"],
+        ),
+        Spacer(1, 8),
+    ]
+    data = [["Control", "Result"]] + [
+        ["Database integrity", qa["database"]["integrity"]],
+        ["Foreign-key violations", qa["database"]["foreign_key_violations"]],
+        ["Workbook status", qa["workbook"]["status"]],
+        ["Workbook sheets", qa["workbook"]["current_sheet_count"]],
+        ["Application status", qa["application"]["status"]],
+        ["Application tests", qa["application"]["test_count"]],
+        ["Publication pages/searchable", f"{qa['publication']['publication_pages']}/{qa['publication']['searchable_pages']}"],
+        ["Publication unchanged", qa["publication"]["publication_unchanged"]],
+        ["Bit Index integrity", qa["indexes"]["bit_index_integrity"]],
+    ]
+    table = Table(data, colWidths=[4.5 * inch, 2.2 * inch], repeatRows=1)
+    table.setStyle(TableStyle([
+        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17324F")),
+        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
+        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
+        ("VALIGN", (0, 0), (-1, -1), "TOP"),
+        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF4F4")]),
+        ("FONTSIZE", (0, 0), (-1, -1), 8),
+    ]))
+    story.extend([
+        table, Spacer(1, 10),
+        Paragraph("Emission policy", styles["Heading1"]),
+        Paragraph(
+            "Full restores are emitted at each session and section boundary and at project completion. Intermediate turns emit checkpoint recovery data. Google Drive is the controlling user-download host.",
+            styles["SmallHP"],
+        ),
+        Spacer(1, 8),
+        Paragraph("Next checkpoint", styles["Heading1"]),
+        Paragraph(
+            "Checkpoint 3 of 3 completes Session 1 and emits the next complete self-contained restore.",
+            styles["SmallHP"],
+        ),
+    ])
+    SimpleDocTemplate(
+        str(path), pagesize=letter, leftMargin=0.65 * inch, rightMargin=0.65 * inch,
+        topMargin=0.6 * inch, bottomMargin=0.6 * inch,
+    ).build(story)
+
+
+def build_report_xlsx(path: Path, qa: dict[str, Any], overlay_rows: list[dict[str, Any]] | None = None) -> None:
+    from openpyxl import Workbook
+    from openpyxl.styles import Alignment, Font, PatternFill
+    from openpyxl.utils import get_column_letter
+
+    wb = Workbook()
+    wb.remove(wb.active)
+    navy, teal, white = "17324F", "167D86", "FFFFFF"
+
+    overview = wb.create_sheet("Overview")
+    overview.append(["Control", "Result", "Evidence"])
+    overview_rows = [
+        ["Checkpoint", "2 of 3", "Section 4 Session 1"],
+        ["Response", 65, "Response_65_Tracking.json"],
+        ["Database integrity", qa["database"]["integrity"], qa["database"]["canonical_database"]],
+        ["Foreign-key violations", qa["database"]["foreign_key_violations"], "DATABASE_QA.json"],
+        ["Workbook status", qa["workbook"]["status"], qa["workbook"]["current_workbook"]],
+        ["Workbook sheets", qa["workbook"]["current_sheet_count"], "Accepted sheets preserved"],
+        ["Application status", qa["application"]["status"], qa["application"]["application"]],
+        ["Application tests", qa["application"]["test_count"], "All return codes zero"],
+        ["Publication unchanged", qa["publication"]["publication_unchanged"], qa["publication"]["publication_sha256"]],
+        ["Bit Index", qa["indexes"]["bit_index_integrity"], qa["indexes"]["bit_index"]],
+    ]
+    for row in overview_rows:
+        overview.append(row)
+
+    for title, payload in [
+        ("Database QA", qa["database"]),
+        ("Workbook QA", qa["workbook"]),
+        ("Application QA", qa["application"]),
+        ("Publication QA", qa["publication"]),
+        ("Index QA", qa["indexes"]),
+    ]:
+        ws = wb.create_sheet(title)
+        ws.append(["Field", "Value"])
+        for key, value in payload.items():
+            ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
+
+    policy = wb.create_sheet("Emission Policy")
+    policy.append(["Boundary", "Required output"])
+    for row in [
+        ["Intermediate checkpoint", "Checkpoint recovery data"],
+        ["Session end", "Complete self-contained restore"],
+        ["Section end", "Complete self-contained restore"],
+        ["Project completion", "Entire final project"],
+        ["Controlling host", "Google Drive"],
+    ]:
+        policy.append(row)
+
+    if overlay_rows is not None:
+        overlay = wb.create_sheet("Overlay Manifest")
+        headers = ["path", "bytes", "sha256", "state"]
+        overlay.append(headers)
+        for row in overlay_rows:
+            overlay.append([row.get(h) for h in headers])
+
+    for ws in wb.worksheets:
+        for cell in ws[1]:
+            cell.font = Font(bold=True, color=white)
+            cell.fill = PatternFill("solid", fgColor=navy)
+        ws.freeze_panes = "A2"
+        ws.auto_filter.ref = ws.dimensions
+        for column in ws.columns:
+            idx = column[0].column
+            width = min(80, max(12, max(len(str(cell.value or "")) for cell in column[:250]) + 2))
+            ws.column_dimensions[get_column_letter(idx)].width = width
+        for row in ws.iter_rows():
+            for cell in row:
+                cell.alignment = Alignment(wrap_text=True, vertical="top")
+    wb.save(path)
+
+
+def compare_overlay(immutable_root: Path, mutable_root: Path) -> tuple[list[dict[str, Any]], list[str]]:
+    immutable_files = {p.relative_to(immutable_root).as_posix(): p for p in immutable_root.rglob("*") if p.is_file()}
+    mutable_files = {p.relative_to(mutable_root).as_posix(): p for p in mutable_root.rglob("*") if p.is_file()}
+    rows = []
+    for rel, path in sorted(mutable_files.items()):
+        old = immutable_files.get(rel)
+        if old and old.stat().st_size == path.stat().st_size and sha256_file(old) == sha256_file(path):
+            continue
+        rows.append({
+            "path": rel,
+            "bytes": path.stat().st_size,
+            "sha256": sha256_file(path),
+            "state": "modified" if old else "new",
+        })
+    deleted = sorted(set(immutable_files) - set(mutable_files))
+    return rows, deleted
+
+
+def build_apply_script(project_root_name: str, overlay_rows: list[dict[str, Any]], deleted: list[str]) -> str:
+    rows_json = json.dumps(overlay_rows, indent=2, ensure_ascii=False)
+    deleted_json = json.dumps(deleted, indent=2, ensure_ascii=False)
+    return f'''#!/usr/bin/env python3
+from __future__ import annotations
+import argparse, hashlib, json, os, shutil, sqlite3, subprocess, sys, tempfile, zipfile
+from pathlib import Path, PurePosixPath
+
+BASE_BYTES = {BASE_RESPONSE64_BYTES}
+BASE_SHA256 = {BASE_RESPONSE64_SHA256!r}
+FINAL_BYTES = {FINAL_SECTION3_BYTES}
+FINAL_SHA256 = {FINAL_SECTION3_SHA256!r}
+PUBLICATION_SHA256 = {PUBLICATION_SHA256!r}
+PROJECT_ROOT_NAME = {project_root_name!r}
+OVERLAY_ROWS = {rows_json}
+DELETED_PATHS = {deleted_json}
+
+def sha256_file(path):
+    h=hashlib.sha256()
+    with open(path,'rb') as f:
+        for b in iter(lambda:f.read(1024*1024),b''):
+            h.update(b)
+    return h.hexdigest()
+
+def safe_extract(zf, target):
+    names=zf.namelist()
+    if len(names)!=len(set(names)):
+        raise SystemExit('Duplicate ZIP member names')
+    for name in names:
+        p=PurePosixPath(name.replace('\\\\','/'))
+        if p.is_absolute() or '..' in p.parts:
+            raise SystemExit(f'Unsafe path: {{name}}')
+    zf.extractall(target)
+
+def locate_base(root):
+    for p in root.rglob('*.zip'):
+        if p.stat().st_size==BASE_BYTES and sha256_file(p)==BASE_SHA256:
+            return p
+    raise SystemExit('Required Complete Restore Through Response 64 not found or identity mismatch')
+
+parser=argparse.ArgumentParser()
+parser.add_argument('--base-response64-restore', required=True, type=Path)
+parser.add_argument('--output-dir', required=True, type=Path)
+parser.add_argument('--output-zip', type=Path)
+parser.add_argument('--skip-tests', action='store_true')
+args=parser.parse_args()
+
+base_arg=args.base_response64_restore
+base=locate_base(base_arg if base_arg.is_dir() else base_arg.parent)
+if base_arg.is_file() and base_arg != base:
+    if base_arg.stat().st_size==BASE_BYTES and sha256_file(base_arg)==BASE_SHA256:
+        base=base_arg
+
+with tempfile.TemporaryDirectory(prefix='mrhpd-cp2-restore-') as td:
+    temp=Path(td)
+    restore=temp/'response64'
+    with zipfile.ZipFile(base) as zf: safe_extract(zf,restore)
+    finals=[p for p in restore.rglob('*FINAL SECTION 3 RELEASE*.zip') if p.stat().st_size==FINAL_BYTES and sha256_file(p)==FINAL_SHA256]
+    if len(finals)!=1: raise SystemExit(f'Final Section 3 release identity failure: {{finals}}')
+    project_extract=temp/'project'
+    with zipfile.ZipFile(finals[0]) as zf: safe_extract(zf,project_extract)
+    dirs=[p for p in project_extract.iterdir() if p.is_dir()]
+    source=dirs[0] if len(dirs)==1 else project_extract
+    output=args.output_dir.resolve()
+    if output.exists(): shutil.rmtree(output)
+    shutil.copytree(source,output)
+    overlay=Path(__file__).resolve().parent/'OVERLAY'/PROJECT_ROOT_NAME
+    for row in OVERLAY_ROWS:
+        src=overlay/row['path']; dst=output/row['path']
+        if not src.exists() or src.stat().st_size!=row['bytes'] or sha256_file(src)!=row['sha256']:
+            raise SystemExit(f'Overlay verification failed: {{row["path"]}}')
+        dst.parent.mkdir(parents=True,exist_ok=True)
+        shutil.copy2(src,dst)
+    for rel in DELETED_PATHS:
+        p=output/rel
+        if p.exists(): p.unlink()
+    dbs=sorted((output/'Database').glob('*Section 4 of 5*Checkpoint 2 of 3*.sqlite'))
+    if len(dbs)!=1: raise SystemExit(f'Canonical database missing: {{dbs}}')
+    con=sqlite3.connect(dbs[0]); integrity=con.execute('PRAGMA integrity_check').fetchone()[0]; fk=list(con.execute('PRAGMA foreign_key_check')); con.close()
+    if integrity!='ok' or fk: raise SystemExit(f'Database QA failed: {{integrity}}, {{fk[:10]}}')
+    pubs=sorted(output.rglob('*Integrated Manuscript*.pdf'),key=lambda p:p.stat().st_size,reverse=True)
+    if not pubs or sha256_file(pubs[0])!=PUBLICATION_SHA256: raise SystemExit('Publication identity changed')
+    if not args.skip_tests:
+        apps=list(output.rglob('human_pathogen_app.py'))
+        if len(apps)!=1: raise SystemExit(f'Application missing: {{apps}}')
+        for test in sorted(apps[0].parent.glob('test*.py')):
+            result=subprocess.run([sys.executable,str(test)],cwd=apps[0].parent,text=True,capture_output=True,timeout=300)
+            if result.returncode: raise SystemExit(f'Application test failed {{test.name}}: {{result.stderr[-2000:]}}')
+    if args.output_zip:
+        args.output_zip.parent.mkdir(parents=True,exist_ok=True)
+        with zipfile.ZipFile(args.output_zip,'w',compression=zipfile.ZIP_DEFLATED,compresslevel=6,allowZip64=True) as zf:
+            for p in sorted(output.rglob('*')):
+                if p.is_file(): zf.write(p,f'{{output.name}}/{{p.relative_to(output).as_posix()}}')
+    print(json.dumps({{'status':'passed','output_dir':str(output),'database_integrity':integrity,'foreign_key_violations':len(fk),'overlay_files':len(OVERLAY_ROWS)}},indent=2))
+'''
+
+
+def package_recovery(immutable_root: Path, mutable_root: Path, dist: Path, qa: dict[str, Any]) -> tuple[Path, dict[str, Any]]:
+    overlay_rows, deleted = compare_overlay(immutable_root, mutable_root)
+    package_root = dist / "checkpoint2_package"
+    if package_root.exists():
+        shutil.rmtree(package_root)
+    package_root.mkdir(parents=True)
+    overlay_root = package_root / "OVERLAY" / mutable_root.name
+    for row in overlay_rows:
+        src = mutable_root / row["path"]
+        dst = overlay_root / row["path"]
+        dst.parent.mkdir(parents=True, exist_ok=True)
+        shutil.copy2(src, dst)
+
+    baseline = {
+        "schema": "mrhpd-checkpoint-recovery-baseline-1.0",
+        "base_complete_restore": {
+            "name_pattern": "*COMPLETE RESTORE THROUGH RESPONSE 64*.zip",
+            "bytes": BASE_RESPONSE64_BYTES,
+            "sha256": BASE_RESPONSE64_SHA256,
+            "google_drive_delivery_folder": "https://drive.google.com/drive/folders/1wTfk19CQA5znGPDK5IIsxxu80dJ57Hdl",
+            "volume_1": "https://drive.google.com/file/d/1szd4sUQu3MIpEfCC4iQCGxJdRl6HlJTt/view?usp=drivesdk",
+            "volume_2": "https://drive.google.com/file/d/1EKvAuw5zEptu9pj_qkQDz_f5HhjKx0Dx/view?usp=drivesdk",
+        },
+        "frozen_final_section3_release": {"bytes": FINAL_SECTION3_BYTES, "sha256": FINAL_SECTION3_SHA256},
+        "accepted_predecessor_mutated": False,
+        "checkpoint": "Remediation Section 4 of 5 Session 1 of 3 Checkpoint 2 of 3",
+        "response": 65,
+    }
+    json_write(package_root / "BASELINE_IDENTITY.json", baseline)
+    manifest = {
+        "schema": "mrhpd-checkpoint-recovery-overlay-1.0",
+        "generated_at": NOW,
+        "project_root_name": mutable_root.name,
+        "base_response64_restore": baseline["base_complete_restore"],
+        "overlay_file_count": len(overlay_rows),
+        "overlay_bytes": sum(row["bytes"] for row in overlay_rows),
+        "deleted_paths": deleted,
+        "files": overlay_rows,
+        "qa": qa,
+        "full_restore_due": False,
+        "full_restore_next_due": "End of Section 4 Session 1 at Checkpoint 3 of 3",
+    }
+    json_write(package_root / "CHECKPOINT_RECOVERY_MANIFEST.json", manifest)
+    text_write(package_root / "DELETED_PATHS.json", json.dumps(deleted, indent=2))
+    text_write(package_root / "RESTORE_READ_FIRST.md", f"""# Human Pathogen Database — Checkpoint 2 Recovery Data Through Response 65
+
+This is an intermediate checkpoint recovery package, not a full project emission. The most recent complete restore is Response 64. This package contains every changed or new project file needed to advance that restore to Section 4 Session 1 Checkpoint 2.
+
+## Required base identity
+
+- Bytes: {BASE_RESPONSE64_BYTES:,}
+- SHA-256: `{BASE_RESPONSE64_SHA256}`
+- Drive folder: https://drive.google.com/drive/folders/1wTfk19CQA5znGPDK5IIsxxu80dJ57Hdl
+
+## Recovery
+
+Run:
+
+```bash
+python TOOLS/apply_checkpoint_recovery.py \\
+  --base-response64-restore "<Complete Restore Through Response 64.zip>" \\
+  --output-dir "<restored Section 4 Checkpoint 2 project>"
+```
+
+The utility verifies the base restore, frozen Section 3 release, every overlay file, SQLite integrity, publication identity and application regressions before reporting success.
+
+## Current state
+
+- Response 65 synchronized
+- Fractional Raw Prompts 64.1–64.3 synchronized
+- Recovery Events 85–90 synchronized
+- Canonical Section 4 database synchronized
+- Comprehensive workbook synchronized without deleting accepted sheets
+- Local application synchronized and regression tested
+- 537-page publication and editable assembly unchanged
+- Next: Checkpoint 3 of 3; end of Session 1 complete restore emission
+""")
+    tools = package_root / "TOOLS"
+    tools.mkdir()
+    text_write(tools / "apply_checkpoint_recovery.py", build_apply_script(mutable_root.name, overlay_rows, deleted))
+
+    reports = package_root / "REPORTS"
+    reports.mkdir()
+    build_report_docx(reports / "MRHPD v3.0.0a Section 4 Session 1 Checkpoint 2 Recovery Report.docx", qa)
+    build_report_pdf(reports / "MRHPD v3.0.0a Section 4 Session 1 Checkpoint 2 Recovery Report.pdf", qa)
+    build_report_xlsx(reports / "MRHPD v3.0.0a Section 4 Session 1 Checkpoint 2 Recovery Register.xlsx", qa, overlay_rows)
+
+    # Package-level checksum inventory.
+    package_files = []
+    for path in sorted(package_root.rglob("*")):
+        if path.is_file() and path.name != "PACKAGE_CHECKSUMS.sha256":
+            package_files.append({"path": path.relative_to(package_root).as_posix(), "bytes": path.stat().st_size, "sha256": sha256_file(path)})
+    text_write(package_root / "PACKAGE_CHECKSUMS.sha256", "".join(f"{row['sha256']}  {row['path']}\n" for row in package_files))
+    json_write(package_root / "PACKAGE_MANIFEST.json", {
+        "generated_at": NOW,
+        "file_count": len(package_files),
+        "total_bytes": sum(row["bytes"] for row in package_files),
+        "files": package_files,
+    })
+
+    archive = dist / (
+        f"Medical References - Human Pathogen Database v{PROJECT_VERSION} Part 8 of 8 "
+        f"Remediation Section 4 of 5 Session 1 of 3 Checkpoint 2 of 3 "
+        f"RECOVERY DATA THROUGH RESPONSE 65 {STAMP}.zip"
+    )
+    with zipfile.ZipFile(archive, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6, allowZip64=True) as zf:
+        for path in sorted(package_root.rglob("*")):
+            if path.is_file():
+                zf.write(path, path.relative_to(package_root).as_posix())
+    archive_qa = verify_zip(archive)
+
+    with tempfile.TemporaryDirectory(prefix="mrhpd-cp2-clean-") as td:
+        clean = Path(td)
+        safe_extract(archive, clean)
+        package_manifest = json.loads((clean / "PACKAGE_MANIFEST.json").read_text(encoding="utf-8"))
+        mismatches = []
+        for row in package_manifest["files"]:
+            p = clean / row["path"]
+            if not p.exists() or p.stat().st_size != row["bytes"] or sha256_file(p) != row["sha256"]:
+                mismatches.append(row["path"])
+        if mismatches:
+            raise RuntimeError({"clean_package_mismatches": mismatches[:30]})
+    verification = {
+        "schema": "mrhpd-checkpoint2-recovery-verification-1.0",
+        "generated_at": NOW,
+        "status": "passed",
+        "archive": archive_qa,
+        "overlay_file_count": len(overlay_rows),
+        "overlay_bytes": sum(row["bytes"] for row in overlay_rows),
+        "deleted_paths": deleted,
+        "package_manifest_mismatches": 0,
+        "base_restore": baseline["base_complete_restore"],
+        "qa": qa,
+        "accepted_predecessor_mutated": False,
+        "checkpoint_2_of_3_complete": True,
+        "session_1_of_3_complete": False,
+        "next_checkpoint": "Checkpoint 3 of 3; complete session restore emission",
+    }
+    json_write(dist / "MRHPD v3.0.0a Response 65 Checkpoint 2 Recovery Verification.json", verification)
+    text_write(dist / f"{archive.name}.sha256.txt", f"{archive_qa['sha256']}  {archive.name}")
+    return archive, verification
+
+
+def main() -> None:
+    parser = argparse.ArgumentParser()
+    parser.add_argument("--base-dir", type=Path, default=Path(os.environ.get("MRHPD_RESPONSE64_RESTORE_DIR", "base_r64")))
+    parser.add_argument("--dist", type=Path, default=Path("dist_cp4_s1_cp2"))
+    args = parser.parse_args()
+    dist = args.dist
+    if dist.exists():
+        shutil.rmtree(dist)
+    dist.mkdir(parents=True)
+
+    base_restore = locate_response64_restore(args.base_dir)
+    base_qa = verify_zip(base_restore, BASE_RESPONSE64_BYTES, BASE_RESPONSE64_SHA256)
+
+    with tempfile.TemporaryDirectory(prefix="mrhpd-cp4-s1-cp2-") as td:
+        work = Path(td)
+        immutable_root, mutable_root, source_qa = extract_final_project(base_restore, work)
+        response64, base_events = load_response64_progress(work / "response64_restore")
+        db, database_qa, events = apply_database_sync(mutable_root, response64, base_events)
+        workbook, workbook_qa = update_workbook(mutable_root, db, database_qa, events)
+        app_files, application_qa = update_application(mutable_root, db)
+        publication_qa = verify_publication(mutable_root)
+        tracking_files = build_tracking_and_governance(
+            mutable_root, database_qa, workbook_qa, application_qa, publication_qa, events
+        )
+        index_qa = build_project_indexes_and_manifest(mutable_root)
+        update_checkpoint_database_status(db, workbook_qa, application_qa, index_qa)
+
+        # Re-run final database verification after all status records are synchronized.
+        con = sqlite3.connect(db)
+        integrity = con.execute("PRAGMA integrity_check").fetchone()[0]
+        fk = list(con.execute("PRAGMA foreign_key_check"))
+        checkpoint_state = con.execute(
+            "SELECT state,workbook_status,application_status FROM section4_checkpoint WHERE checkpoint_code='MRHPD-V3-CP4-S1-CP2'"
+        ).fetchone()
+        con.close()
+        if integrity != "ok" or fk or not checkpoint_state or checkpoint_state[0] != "checkpoint_complete":
+            raise RuntimeError({"final_database_gate": {"integrity": integrity, "fk": fk[:20], "checkpoint_state": checkpoint_state}})
+        database_qa["final_integrity"] = integrity
+        database_qa["final_foreign_key_violations"] = len(fk)
+        database_qa["checkpoint_state"] = list(checkpoint_state)
+        database_qa["sha256_after_final_status_sync"] = sha256_file(db)
+
+        qa = {
+            "source": source_qa,
+            "base_restore": base_qa,
+            "database": database_qa,
+            "workbook": workbook_qa,
+            "application": application_qa,
+            "publication": publication_qa,
+            "indexes": index_qa,
+            "tracking_file_count": len(tracking_files),
+            "accepted_predecessor_mutated": False,
+        }
+        qa_dir = mutable_root / "QA" / "Section 4 Session 1" / "Checkpoint 2"
+        json_write(qa_dir / "CHECKPOINT_2_FINAL_QA.json", qa)
+        archive, verification = package_recovery(immutable_root, mutable_root, dist, qa)
+        json_write(dist / "MRHPD_RESPONSE65_CHECKPOINT2_BUILD_SUMMARY.json", {
+            "status": "passed",
+            "generated_at": NOW,
+            "archive": {"file": archive.name, "bytes": archive.stat().st_size, "sha256": sha256_file(archive)},
+            "checkpoint": "Remediation Section 4 of 5 Session 1 of 3 Checkpoint 2 of 3 COMPLETE",
+            "session_complete": False,
+            "full_restore_due_this_checkpoint": False,
+            "next": "Checkpoint 3 of 3; complete Session 1 restore emission",
+        })
+        print(json.dumps(verification, indent=2, ensure_ascii=False))
+
+
+if __name__ == "__main__":
+    main()
